import itertools
import time
import warnings
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple, Union

import numpy as np
import pandas
from pandas.errors import PerformanceWarning

BUCKET_SCALES = [-np.inf, -20, -10, -5, -2, 0, 2, 5, 10, 20, np.inf]
BUCKETS = [
    "<-20%",
    "[-20%,-10%[",
    "[-10%,-5%[",
    "[-5%,-2%[",
    "[-2%,0%[",
    "[0%,2%[",
    "[2%,5%[",
    "[5%,10%[",
    "[10%,20%[",
    ">=20%",
    "script <-20%",
    "script [-20%,-10%[",
    "script [-10%,-5%[",
    "script [-5%,-2%[",
    "script [-2%,0%[",
    "script [0%,2%[",
    "script [2%,5%[",
    "script [5%,10%[",
    "script [10%,20%[",
    "script >=20%",
]


def _SELECTED_FEATURES():
    features = [
        dict(
            cat="status",
            stat="date",
            agg="MAX",
            new_name="date",
            unit="date",
            help="Most recent date involved",
            simple=True,
        ),
        dict(
            cat="time",
            stat="ITER",
            agg="SUM",
            new_name="benchmark duration",
            unit="s",
            help="Total duration of the benchmark",
            simple=True,
        ),
        dict(
            cat="time",
            stat="ITER",
            agg="TOTAL",
            new_name="number of models",
            unit="N",
            help="Number of models evaluated in this document.",
            simple=True,
        ),
        dict(
            cat="time",
            stat="latency_eager",
            agg="COUNT",
            new_name="number of models eager mode",
            unit="N",
            help="Number of models running with eager mode",
            simple=True,
        ),
        dict(
            cat="status",
            stat="control_flow",
            agg="SUM",
            new_name="number of failures for torch.export.export",
            unit="N",
            help="torch.export.export does not work because of a "
            "control flow or any other reason, in practice, "
            "this column means the number of models torch.export.export "
            "fails, this metric is only available if the data of exporter "
            "'export' or 'compile' is aggregated",
            simple=True,
        ),
        dict(
            cat="time",
            agg="COUNT",
            stat="export_unbiased",
            new_name="export number",
            unit="N",
            help="Number of models successfully converted into ONNX. "
            "The ONNX model may not be run through onnxruntime or with "
            "significant discrepancies.",
            simple=True,
        ),
        dict(
            cat="speedup",
            agg="COUNT",
            stat="increase",
            new_name="number of running models",
            unit="N",
            help="Number of models converted and running with onnxruntime. "
            "The outputs may be right or wrong. Unit test ensures every aten functions "
            "is correctly converted but the combination may produce outputs "
            "with higher discrepancies than expected.",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="accuracy_rate",
            new_name="accuracy number",
            unit="N",
            help="Number of models successfully converted into ONNX. "
            "It may be slow but the discrepancies are < 0.1.",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="accuracy_dynamic_rate",
            new_name="accuracy number for different dynamic shapes",
            unit="N",
            help="Number of models successfully converted into ONNX. "
            "It may be slow but the discrepancies are < 0.1 for a different set "
            "of inputs with different shapes",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="pass_rate",
            new_name="pass number",
            unit="N",
            help="Number of models successfully converted into ONNX, "
            "with a maximum discrepancy < 0.1 and a speedup > 0.98.",
            simple=True,
        ),
        dict(
            cat="time",
            agg="SUM",
            stat="export_unbiased",
            new_name="total export time",
            unit="s",
            help="Total export time when the export succeeds. "
            "The model may not run through onnxruntime and the model "
            "may produce higher discrepancies than expected (lower is better).",
            simple=True,
        ),
        dict(
            cat="time",
            agg="SUM",
            stat="latency",
            new_name="total latency time exported model",
            unit="x",
            help="Total latency time with the exported model (onnxruntime, inductor, ...)",
            simple=True,
        ),
        dict(
            cat="time",
            agg="SUM",
            stat="latency_eager_if_exported_run",
            new_name="total latency time eager / exported model",
            unit="x",
            help="Total latency of eager mode knowing when the "
            "exported model (onnxruntime, inductor, ...) runs",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="lat<=script+2%",
            new_name="number of models equal or faster than torch.script",
            unit="N",
            help="Number of models successfully converted with torch.script "
            "and the other exporter, and the second exporter is as fast or faster "
            "than torch.script.",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="lat<=eager+2%",
            new_name="number of models equal or faster than eager",
            unit="N",
            help="Number of models as fast or faster than torch eager mode.",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="lat<=inductor+2%",
            new_name="number of models equal or faster than inductor",
            unit="N",
            help="Number of models equal or faster than inductor (fullgraph=True)",
            simple=True,
        ),
        # average
        dict(
            cat="time",
            agg="MEAN",
            stat="export_unbiased",
            new_name="average export time",
            unit="s",
            help="Average export time when the export succeeds. "
            "The model may not run through onnxruntime and the model "
            "may produce higher discrepancies than expected (lower is better).",
            simple=True,
        ),
        dict(
            cat="speedup",
            agg="GEO-MEAN",
            stat="1speedup",
            new_name="average speedup (geo)",
            unit="x",
            help="Geometric mean of all speedup for all model converted and runnning.",
            simple=True,
        ),
        # e-1
        dict(
            cat="status",
            agg="SUM",
            stat="err<1e-1",
            new_name="discrepancies < 0.1",
            unit="N",
            help="Number of models for which the maximum discrepancies is "
            "below 0.1 for all outputs.",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="dynerr<1e-1",
            new_name="dynamic discrepancies < 0.1",
            unit="N",
            help="Number of models for which the maximum discrepancies "
            "with other dynamic shapes is below 0.1 for all outputs.",
            simple=True,
        ),
        # e-2
        dict(
            cat="status",
            agg="SUM",
            stat="err<1e-2",
            new_name="discrepancies < 0.01",
            unit="N",
            help="Number of models for which the maximum discrepancies is "
            "below 0.01 for all outputs.",
            simple=True,
        ),
        dict(
            cat="time",
            stat="ITER",
            agg="MEAN",
            new_name="average iteration time",
            unit="s",
            help="Average total time per model and scenario. "
            "It usually reflects how long the export time is (lower is better).",
        ),
        dict(
            cat="discrepancies",
            stat="avg",
            agg="MEAN",
            new_name="average average discrepancies",
            unit="f",
            help="Average of average absolute discrepancies "
            "assuming it can be measured (lower is better).",
        ),
        dict(
            cat="discrepancies",
            stat="abs",
            agg="MEAN",
            new_name="average absolute discrepancies",
            unit="f",
            help="Average maximum absolute discrepancies "
            "assuming it can be measured (lower is better).",
        ),
        dict(
            cat="time",
            agg="MEAN",
            stat="latency_eager",
            new_name="average latency eager",
            unit="s",
            help="Average latency for eager mode (lower is better)",
        ),
        dict(
            cat="time",
            agg="MEAN",
            stat="latency",
            new_name="average latency ort",
            unit="s",
            help="Average latency for onnxruntime (lower is better)",
        ),
        dict(
            cat="onnx",
            agg="MEAN",
            stat="weight_size_torch",
            new_name="average weight size",
            unit="bytes",
            help="Average parameters size, this gives a kind of order "
            "of magnitude for the memory peak",
        ),
        dict(
            cat="onnx",
            agg="MEAN",
            stat="weight_size_proto",
            new_name="average weight size in ModelProto",
            unit="bytes",
            help="Average parameters size in the model proto, "
            "this gives a kind of order of magnitude for the memory peak "
            "this should be close to the parameter size",
        ),
        dict(
            cat="onnx",
            agg="MAX",
            stat="weight_size_torch",
            new_name="maximum weight size",
            unit="bytes",
            help="Maximum parameters size, "
            "useful to guess how much this machine can handle",
        ),
        dict(
            cat="onnx",
            agg="MAX",
            stat="weight_size_proto",
            new_name="maximum weight size in ModelProto",
            unit="bytes",
            help="Maximum parameters size in the model proto, "
            "useful to guess how much this machine can handle",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="peak_gpu_eager_warmup",
            new_name="average GPU peak (eager warmup)",
            unit="bytes",
            help="Average GPU peak while warming up eager mode (torch metric)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_gpu_eager_warmup",
            new_name="average GPU delta peak (eager warmup)",
            unit="bytes",
            help="Average GPU peak of new allocated memory while warming up eager "
            "mode (torch metric)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="peak_gpu_warmup",
            new_name="average GPU peak (warmup)",
            unit="bytes",
            help="Average GPU peak while warming up onnxruntime (torch metric)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_gpu_warmup",
            new_name="average GPU delta peak (warmup)",
            unit="bytes",
            help="Average GPU peak of new allocated memory while warming up "
            "onnxruntime (torch metric)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_cpu_pp",
            new_name="average CPU delta peak (export)",
            unit="bytes",
            help="Average CPU peak of new allocated memory while converting "
            "the model (measured in a secondary process)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_gpu_export",
            new_name="average GPU delta peak (export) (torch)",
            unit="bytes",
            help="Average GPU peak of new allocated memory while "
            "converting the model (torch metric)",
            simple=True,
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_gpu_pp",
            new_name="average GPU delta peak (export) (nvidia-smi)",
            unit="bytes",
            help="Average GPU peak of new allocated memory while "
            "converting the model (measured in a secondary process)",
            simple=True,
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="peak_gpu_export",
            new_name="average GPU peak (export)",
            unit="bytes",
            help="Average GPU peak while converting the model (torch metric)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_gpu_export",
            new_name="average delta GPU peak (export)",
            unit="bytes",
            help="Average GPU peak of new allocated memory "
            "while converting the model (torch metric)",
            simple=True,
        ),
        dict(
            cat="speedup",
            agg="MEAN",
            stat="increase",
            new_name="average speedup increase",
            unit="%",
            help="Average speedup increase compare to eager mode.",
        ),
    ]
    for b in BUCKETS:
        s = "script" in b
        bs = b.replace("script ", "")
        ag = "torch_script" if s else "eager"
        features.append(
            dict(
                cat="bucket",
                agg="SUM",
                stat=b,
                new_name=f"speedup/script in {bs}" if s else f"speedup in {bs}",
                unit="N",
                help=f"Number of models whose speedup against {ag} "
                f"falls into this interval",
                simple=True,
            )
        )
    return features


SELECTED_FEATURES = _SELECTED_FEATURES()

FILTERS = {
    "HG": [
        "AlbertForMaskedLM",
        "AlbertForQuestionAnswering",
        "AllenaiLongformerBase",
        "BartForCausalLM",
        "BartForConditionalGeneration",
        "BertForMaskedLM",
        "BertForQuestionAnswering",
        "BlenderbotForCausalLM",
        "BlenderbotSmallForCausalLM",
        "BlenderbotSmallForConditionalGeneration",
        "CamemBert",
        "DebertaForMaskedLM",
        "DebertaForQuestionAnswering",
        "DebertaV2ForMaskedLM",
        "DebertaV2ForQuestionAnswering",
        "DistilBertForMaskedLM",
        "DistilBertForQuestionAnswering",
        "DistillGPT2",
        "ElectraForCausalLM",
        "ElectraForQuestionAnswering",
        "GPT2ForSequenceClassification",
        "GoogleFnet",
        "LayoutLMForMaskedLM",
        "LayoutLMForSequenceClassification",
        "M2M100ForConditionalGeneration",
        "MBartForCausalLM",
        "MBartForConditionalGeneration",
        "MT5ForConditionalGeneration",
        "MegatronBertForCausalLM",
        "MegatronBertForQuestionAnswering",
        "MobileBertForMaskedLM",
        "MobileBertForQuestionAnswering",
        "OPTForCausalLM",
        "PLBartForCausalLM",
        "PLBartForConditionalGeneration",
        "PegasusForCausalLM",
        "PegasusForConditionalGeneration",
        "RobertaForCausalLM",
        "RobertaForQuestionAnswering",
        "Speech2Text2ForCausalLM",
        "T5ForConditionalGeneration",
        "T5Small",
        "TrOCRForCausalLM",
        "XGLMForCausalLM",
        "XLNetLMHeadModel",
        "YituTechConvBert",
    ]
}


def _nonone_(index):
    return None not in index.names


def _key(v):
    if isinstance(v, (int, float)):
        return v, ""
    if isinstance(v, str):
        return 1e10, v
    if isinstance(v, tuple):
        return (1e10, *v)
    raise AssertionError(f"Unexpected type for v={v!r}, type is {type(v)}")


def sort_index_key(index):
    return pandas.Index(_key(v) for v in index)


def _isnan(x):
    if x is None:
        return True
    if isinstance(x, str):
        return x == ""
    try:
        return np.isnan(x)
    except TypeError:
        return False


def _isinf(x):
    if x is None:
        return True
    if isinstance(x, str):
        return x == "inf"
    try:
        return np.isinf(x)
    except TypeError:
        return False


def _column_name(ci: int) -> str:
    if ci < 26:
        return chr(65 + ci)
    a = ci // 26
    b = ci % 26
    return f"{chr(65+a)}{chr(65+b)}"


def _format_excel_cells(
    sheets: Union[str, List[str]],
    writer: pandas.ExcelWriter,
    verbose: int = 0,
):
    if isinstance(sheets, list):
        for sheet in sheets:
            _format_excel_cells(sheet, writer, verbose=verbose)
        return

    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    bold_font = Font(bold=True)
    left = Alignment(horizontal="left")
    right = Alignment(horizontal="right")

    sheet = writer.sheets[sheets]
    rows = sheet.max_row
    co = {}
    sizes = {}
    cols = set()
    for i in range(1, rows + 1):
        for j, cell in enumerate(sheet[i]):
            cols.add(cell.column)
            if isinstance(cell.value, float):
                co[j] = co.get(j, 0) + 1
            elif isinstance(cell.value, str):
                sizes[cell.column] = max(sizes.get(cell.column, 0), len(cell.value))

    for k, v in sizes.items():
        c = get_column_letter(k)
        sheet.column_dimensions[c].width = max(15, v)
    for k in cols:
        if k not in sizes:
            c = get_column_letter(k)
            sheet.column_dimensions[c].width = 15

    top = [(v, k) for k, v in co.items()]
    top.sort()
    i = len(top) - 1
    while i > 0 and top[i - 1][0] == top[-1][0]:
        i -= 1
    first_col = top[i][-1]
    for i in range(1, rows + 1):
        for j, cell in enumerate(sheet[i]):
            if j < first_col:
                cell.font = bold_font
                cell.alignment = left
            elif isinstance(cell.value, float):
                cell.alignment = right
                x = cell.value
                if int(x) == x:
                    cell.number_format = "0"
                elif x > 1000:
                    cell.number_format = "0 000"
                elif x >= 100:
                    cell.number_format = "0.0"
                elif x >= 10:
                    cell.number_format = "0.00"
                elif x >= 1:
                    cell.number_format = "0.000"
                elif x > 0.1:
                    cell.number_format = "0.000"
                else:
                    cell.number_format = "0.000000"


def _apply_excel_style(
    res: Dict[str, pandas.DataFrame],
    writer: pandas.ExcelWriter,
    verbose: int = 0,
):
    from openpyxl.styles import Alignment, Font, PatternFill, numbers

    bold_font = Font(bold=True)
    alignment = Alignment(horizontal="left")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    red = Font(color="FF0000")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    redf = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    lasts = {}
    for k, v in res.items():
        if verbose:
            print(f"[_apply_excel_style] apply style on {k!r}, shape={v.shape}")
        sheet = writer.sheets[k]
        if 0 in v.shape:
            continue

        if k in ("0main", "0main_base"):
            for c in "AB":
                sheet.column_dimensions[c].width = 40
                sheet.column_dimensions[c].alignment = alignment
            for cell in sheet[1]:
                cell.font = bold_font
                cell.alignment = alignment
            continue

        if k in {"0raw", "AGG", "AGG2", "0raw_base"}:
            continue

        n_cols = (
            1 if isinstance(v.index[0], (str, int, np.int64, np.int32)) else len(v.index[0])
        )
        n_rows = (
            1
            if isinstance(v.columns[0], (str, int, np.int64, np.int32))
            else len(v.columns[0])
        )

        for i in range(n_cols):
            sheet.column_dimensions[_column_name(i)].width = 40

        first_row = None
        first_col = None
        if v.index.names == [None]:
            first_row = len(v.columns.names)
            first_col = len(v.columns.names)
            if verbose > 1:
                print(
                    f"[_apply_excel_style] k={k!r}, first={first_row},{first_col}, "
                    f"columns.names={v.columns.names}"
                )
        else:
            look = v.iloc[0, 0]

            values = []
            for row in sheet.iter_rows(
                min_row=1,
                max_row=n_rows + n_cols + 1,
                min_col=1,
                max_col=n_rows + n_cols + 1,
            ):
                for cell in row:
                    if hasattr(cell, "col_idx") and (
                        cell.value == look
                        or (_isnan(cell.value) and _isnan(look))
                        or (_isinf(cell.value) and _isinf(look))
                    ):
                        first_row = cell.row
                        first_col = cell.col_idx if hasattr(cell, "col_idx") else first_col
                        break
                    values.append(cell.value)
                if first_row is not None:
                    break

            if verbose > 1:
                print(
                    f"[_apply_excel_style] k={k!r}, first={first_row},{first_col}, "
                    f"index.names={v.index.names}, columns.names={v.columns.names}"
                )

            assert first_row is not None and first_col is not None, (
                f"Unable to find the first value in {k!r}, first_row={first_row}, "
                f"first_col={first_col}, look={look!r} ({type(look)}), "
                f"n_rows={n_rows}, n_cols={n_cols}, values={values}, "
                f"iloc[:3,:3]={v.iloc[:3, :3]}, v.index.names={v.index.names}, "
                f"v.columns={v.columns}"
            )

        last_row = first_row + v.shape[0] + 1
        last_col = first_col + v.shape[1] + 1
        lasts[k] = (last_row, last_col)
        for row in sheet.iter_rows(
            min_row=first_row,
            max_row=last_row,
            min_col=first_col,
            max_col=last_col,
        ):
            for cell in row:
                cell.alignment = alignment
                cell.font = bold_font
        for i in range(n_rows):
            for cell in sheet[i + 1]:
                cell.font = bold_font

        if k == "memory":
            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    cell.number_format = numbers.FORMAT_NUMBER
            continue

        if k == "ERR":
            for i in range(n_cols + v.shape[1]):
                sheet.column_dimensions[_column_name(i)].alignment = alignment
                sheet.column_dimensions[_column_name(i)].width = 50
                if i >= 25:
                    break
            continue

        if k in ("TIME_ITER", "discrepancies", "speedup"):
            c1, c2 = None, None
            qpb, qpt, fmtp = None, None, None
            if k == "TIME_ITER":
                qt = np.quantile(v.values, 0.9)
                qb = None
                fmt = numbers.FORMAT_NUMBER_00
            elif k == "speedup":
                debug_values = []
                for row in sheet.iter_rows(
                    min_row=0,
                    max_row=30,
                    min_col=first_col,
                    max_col=last_col,
                ):
                    found = None
                    for cell in row:
                        debug_values.append(cell.value)
                        if cell.value == "increase":
                            c1 = cell.col_idx if c1 is None else min(cell.col_idx, c1)
                            c2 = cell.col_idx if c2 is None else max(cell.col_idx, c2)
                            found = cell
                            last_idx = cell.col_idx
                        elif cell.value is None:
                            # merged cell
                            if found is not None:
                                last_idx += 1
                                cid = getattr(cell, "col_idx", last_idx)
                                c2 = max(cid, c2)
                        else:
                            found = None

                if not (c1 is None and c2 is None):
                    assert c1 is not None and c2 is not None and c1 <= c2, (
                        f"Unexpected value for c1={c1}, c2={c2} (k={k!r})"
                        f"\ndebug_values={debug_values}"
                    )
                # ratio
                qb = 0.98
                qt = None
                fmt = "0.0000"
                # increase
                qpb = -0.02
                qpt = None
                fmtp = "0.000%"
            else:
                qt = 0.01
                qb = None
                fmt = "0.000000"

            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    if cell.value is None or isinstance(cell.value, str):
                        continue
                    if c1 is not None and cell.col_idx >= c1 and cell.col_idx <= c2:
                        if qt and cell.value > qpt:
                            cell.font = red
                        if qb and cell.value < qpb:
                            cell.font = red
                        cell.number_format = fmtp
                    else:
                        if qt and cell.value > qt:
                            cell.font = red
                        if qb and cell.value < qb:
                            cell.font = red
                        cell.number_format = fmt
            continue

        if k in ("bucket", "status", "op_onnx", "op_opt", "op_torch") and v.shape[1] > 3:
            has_convert = [("convert" in str(c)) for c in v.columns]
            has_20 = [("-20%" in str(c)) for c in v.columns]
            assert not k.startswith("bucket") or any(
                has_20
            ), f"has_20={has_20} but k={k!r}, v.columns={[str(c) for c in v.columns]}"
            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    if cell.value is None or isinstance(cell.value, str):
                        continue
                    if cell.value == 0:
                        cell.value = None
                    cell.number_format = numbers.FORMAT_NUMBER
                    cell.alignment = center
                    if k == "status":
                        idx = cell.col_idx - n_cols - 1
                        if has_convert[idx]:
                            cell.fill = yellow
                    elif k in ("bucket",):
                        idx = cell.col_idx - n_cols - 1
                        if has_20[idx]:
                            cell.fill = redf
            continue

        if k == "time":
            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    if cell.value is None or isinstance(cell.value, str):
                        continue
                    cell.number_format = "0.000000"
            continue

        if k in {
            "SUMMARY",
            "SUMMARY2",
            "SUMMARY_base",
            "SUMMARY2_base",
            "SUMMARY2_diff",
            "SUMMARY_diff",
            "SIMPLE",
            "SIMPLE_base",
            "SIMPLE_diff",
        }:
            fmt = {
                "x": "0.000",
                "%": "0.000%",
                "bytes": "# ##0",
                "Mb": "0.000",
                "N": "0",
                "f": "0.000",
                "s": "0.0000",
                "date": "aaaa-mm-dd",
            }
            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    if cell.value in fmt:
                        start, end = (
                            (0, cell.col_idx) if "SUMMARY" in k else (cell.col_idx, last_col)
                        )
                        for idx in range(start, end):
                            fcell = row[idx]
                            if isinstance(fcell.value, (int, float)):
                                f = fmt[cell.value]
                                if cell.value == "x":
                                    if (
                                        isinstance(fcell.value, (float, int))
                                        and fcell.value < 0.98
                                    ):
                                        fcell.font = red
                                elif cell.value in ("f", "s", "Mb"):
                                    if fcell.value >= 1000:
                                        f = "0 000"
                                    elif fcell.value >= 10:
                                        f = "0.00"
                                    elif fcell.value >= 1:
                                        f = "0.000"
                                elif cell.value == "date" and "SIMPLE" not in k:
                                    ts = time.gmtime(fcell.value)
                                    sval = time.strftime("%Y-%m-%d", ts)
                                    fcell.value = sval
                                fcell.number_format = f

            cols = {}
            for row in sheet.iter_rows(
                min_row=1,
                max_row=2,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    if isinstance(cell.value, str):
                        cols[cell.value] = cell.col_idx

            maxc = None
            done = set()
            for k, ci in cols.items():
                if k is None or isinstance(k, int):
                    continue
                c = _column_name(ci - 1)
                if k in ("order", "#order"):
                    sheet.column_dimensions[c].width = 5
                    for cell in sheet[c]:
                        cell.alignment = right
                    done.add(c)
                    continue
                if k == "METRIC":
                    sheet.column_dimensions[c].width = 50
                    for cell in sheet[c]:
                        cell.alignment = alignment
                    done.add(c)
                    continue
                if k in {
                    "exporter",
                    "opt_patterns",
                    "dynamic",
                    "rtopt",
                    "suite",
                } or k.startswith("version"):
                    sheet.column_dimensions[c].width = 15
                    for cell in sheet[c]:
                        cell.alignment = alignment
                    done.add(c)
                    continue
                if k == "unit":
                    sheet.column_dimensions[c].width = 7
                    for cell in sheet[c]:
                        cell.alignment = right
                    maxc = c
                    done.add(c)
                    continue
                if k in ("help", "~help"):
                    sheet.column_dimensions[c].width = 20
                    for cell in sheet[c]:
                        cell.alignment = alignment
                    done.add(c)
                    continue

            for c in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                if c == maxc:
                    break
                if c in done:
                    continue
                sheet.column_dimensions[c].width = 20
                for cell in sheet[c]:
                    cell.alignment = right
            continue

    if verbose:
        print("[_apply_excel_style] done")


def _reorder_columns_level(
    df: pandas.DataFrame,
    first_level: List[str],
    prefix: Optional[str] = None,
) -> pandas.DataFrame:
    assert _nonone_(df.columns), f"None in {df.columns.names}, prefix={prefix!r}"
    assert set(df.columns.names) & set(first_level), (
        f"Nothing to sort, prefix={prefix!r} "
        f"df.columns={df.columns}, first_level={first_level}\n--\n{df}"
    )

    c_in = [c for c in first_level if c in set(df.columns.names)]
    c_out = [c for c in df.columns.names if c not in set(c_in)]
    levels = c_in + c_out
    for i in range(len(levels)):
        if levels[i] == df.columns.names[i]:
            continue
        j = list(df.columns.names).index(levels[i])
        df = df.swaplevel(i, j, axis=1)
    assert (
        list(df.columns.names) == levels
    ), f"Issue levels={levels}, df.columns.names={df.columns.names}"
    assert _nonone_(df.columns), f"None in {df.columns.names}"
    return df.sort_index(axis=1)


def _sort_index_level(
    df: pandas.DataFrame,
    debug: Optional[str] = None,
) -> pandas.DataFrame:
    assert _nonone_(df.index), f"None in {df.index.names}, debug={debug!r}"
    assert df.columns.names == [None] or _nonone_(
        df.columns
    ), f"None in {df.columns.names}, debug={debug!r}"
    levels = list(df.index.names)
    levels.sort()
    for i in range(len(levels)):
        if levels[i] == df.index.names[i]:
            continue
        j = list(df.index.names).index(levels[i])
        df = df.swaplevel(i, j, axis=0)
    assert (
        list(df.index.names) == levels
    ), f"Issue levels={levels}, df.index.names={df.index.names}, debug={debug!r}"
    assert _nonone_(df.index), f"None in {df.index.names}, debug={debug!r}"
    assert df.columns.names == [None] or _nonone_(
        df.columns
    ), f"None in {df.columns.names}, debug={debug!r}"
    return df.sort_index(axis=0)


def _reorder_index_level(
    df: pandas.DataFrame,
    first_level: List[str],
    prefix: Optional[str] = None,
) -> pandas.DataFrame:
    assert _nonone_(df.index), f"None in {df.index.names}, prefix={prefix!r}"
    assert _nonone_(df.columns), f"None in {df.columns.names}, prefix={prefix!r}"
    assert set(df.index.names) & set(first_level), (
        f"Nothing to sort, prefix={prefix!r} "
        f"df.columns={df.index}, first_level={first_level}"
    )

    c_in = [c for c in first_level if c in set(df.index.names)]
    c_out = [c for c in df.index.names if c not in set(c_in)]
    levels = c_in + c_out
    for i in range(len(levels)):
        if levels[i] == df.index.names[i]:
            continue
        j = list(df.index.names).index(levels[i])
        df = df.swaplevel(i, j, axis=0)
    assert (
        list(df.index.names) == levels
    ), f"Issue levels={levels}, df.columns.names={df.index.names}"
    assert _nonone_(df.index), f"None in {df.index.names}"
    assert _nonone_(df.columns), f"None in {df.columns.names}"
    return df.sort_index(axis=0)


def _add_level(
    index: pandas.MultiIndex,
    name: str,
    value: str,
) -> pandas.MultiIndex:
    if len(index.names) == 1:
        v = index.tolist()
        nv = [[value, _] for _ in v]
        nn = [name, index.names[0]]
        aa = np.array(nv).T.tolist()
        new_index = pandas.MultiIndex.from_arrays(aa, names=nn)
        return new_index

    v = index.tolist()
    nv = [[value, *_] for _ in v]
    nn = [name, *index.names]
    aa = np.array(nv).T.tolist()
    new_index = pandas.MultiIndex.from_arrays(aa, names=nn)
    return new_index


def _create_aggregation_figures(
    final_res: Dict[str, pandas.DataFrame],
    model: List[str],
    skip: Optional[Set[str]] = None,
    key: str = "suite",
    exc: bool = True,
) -> Dict[str, pandas.DataFrame]:
    assert key in model, f"Key {key!r} missing in model={model!r}"
    model_not_key = [c for c in model if c != key]

    aggs = {}
    for k, v in final_res.items():
        if k in skip:
            continue
        n_numerical = v.select_dtypes(include=np.number).shape[1]
        assert k in {"onnx"} or (
            n_numerical > 0
        ), f"No numeric column for k={k!r}, dtypes=\n{v.dtypes}"
        assert None not in v.index.names, f"None in v.index.names={v.index.names}, k={k!r}"
        assert (
            None not in v.columns.names
        ), f"None in v.columns.names={v.columns.names}, k={k!r}"

        if key not in v.index.names:
            v = v.copy()
            v[key] = "?"
            v = v.reset_index(drop=False).set_index([key, *v.index.names])

        assert key in v.index.names, f"Unable to find key={key} in {v.index.names} for k={k!r}"
        assert len(v.index.names) == len(
            model
        ), f"Length mismatch for k={k!r}, v.index.names={v.index.names}, model={model}"

        # Let's drop any non numerical features.
        v = v.select_dtypes(include=[np.number])
        # gv = v.apply(lambda x: np.log(np.maximum(x, 1e-10).values))
        v = v.reset_index(drop=False).set_index(model_not_key)

        assert key in v.columns, f"Unable to find column {key!r} in {v.columns}"

        v = v.sort_index(axis=1)

        assert None not in v.index.names, f"None in v.index.names={v.index.names}, k={k!r}"
        assert (
            None not in v.columns.names
        ), f"None in v.columns.names={v.columns.names}, k={k!r}"
        try:
            gr = v.groupby(key)
        except ValueError as e:
            raise AssertionError(
                f"Unable to grouby by key={key!r}, "
                f"shape={v.shape} v.columns={v.columns}, "
                f"values={set(v[key])}\n---\n{v}"
            ) from e

        def _geo_mean(serie):
            nonan = serie.dropna()
            if len(nonan) == 0:
                return np.nan
            res = np.exp(np.log(np.maximum(nonan, 1e-10)).mean())
            return res

        def _propnan(df, is_nan):
            df[is_nan] = np.nan
            return df

        gr_no_nan = v.fillna(0).groupby(key)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=(FutureWarning, PerformanceWarning))
            total = gr_no_nan.count()
            is_nan = gr_no_nan.count() - gr.count() == total
        stats = [
            ("MEAN", gr.mean()),
            ("MEDIAN", gr.median()),
            ("SUM", _propnan(gr.sum(), is_nan)),
            ("MIN", gr.min()),
            ("MAX", gr.max()),
            ("COUNT", _propnan(gr.count(), is_nan)),
            ("COUNT%", _propnan(gr.count() / total, is_nan)),
            ("TOTAL", _propnan(total, is_nan)),
            ("NAN", _propnan(is_nan.astype(int), is_nan)),
        ]

        if k.startswith("speedup"):
            try:
                geo_mean = gr.agg(_geo_mean)
            except ValueError as e:
                if exc:
                    raise AssertionError(
                        f"Fails for geo_mean, k={k!r}, v=\n{v.head().T}"
                    ) from e
                geo_mean = None
            if geo_mean is not None:
                stats.append(("GEO-MEAN", geo_mean))

        dfs = []
        for name, df in stats:
            # avoid date to be numbers
            updates = {}
            drops = []
            for col in df.columns:
                if (isinstance(col, tuple) and "date" in col) or col == "date":
                    if name in {"SUM", "MEDIAN"}:
                        drops.append(col)
                    elif name in {"MIN", "MAX", "MEAN", "MEDIAM"}:
                        # maybe this code will fail someday but it seems that the cycle
                        # date -> int64 -> float64 -> int64 -> date
                        # keeps the date unchanged
                        vvv = df[col]
                        if vvv.dtype not in {object, np.object_, str, np.str_}:
                            updates[col] = vvv.apply(
                                lambda d: (
                                    np.nan
                                    if np.isnan(d)
                                    else time.strftime("%Y-%m-%d", time.gmtime(d))
                                )
                            )
            if drops:
                df.drop(drops, axis=1, inplace=True)
            if updates:
                for kc, vc in updates.items():
                    df[kc] = vc

            # then continue
            assert isinstance(
                df, pandas.DataFrame
            ), f"Unexpected type {type(df)} for k={k!r} and name={name!r}"
            df.index = _add_level(df.index, "agg", name)
            df.index = _add_level(df.index, "cat", k)
            assert _nonone_(df.index), f"None in {df.index.names}, k={k!r}, name={name!r}"
            assert _nonone_(df.columns), f"None in {df.columns.names}, k={k!r}, name={name!r}"
            dfs.append(df)

        if len(dfs) == 0:
            continue
        df = pandas.concat(dfs, axis=0)

        if df.shape[1] == 0 and k in {"onnx"}:
            # nothing to do
            continue

        assert df.shape[0] > 0, f"Empty set for k={k!r}"
        assert df.shape[1] > 0, f"Empty columns for k={k!r}"
        assert _nonone_(df.index), f"None in {df.index.names}, k={k!r}"
        assert _nonone_(df.columns), f"None in {df.columns.names}, k={k!r}"
        assert isinstance(df, pandas.DataFrame), f"Unexpected type {type(df)} for k={k!r}"

        if "stat" in df.columns.names:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=(FutureWarning, PerformanceWarning))
                df = df.stack("stat")
            if not isinstance(df, pandas.DataFrame):
                assert (
                    "opt_patterns" not in df.index.names
                    and "rtopt" not in df.index.names
                    and "dynamic" not in df.index.names
                ), f"Unexpected names for df.index.names={df.index.names} (k={k!r})"
                df = df.to_frame()
                if df.columns.names == [None]:
                    df.columns = pandas.MultiIndex.from_arrays(
                        [("_dummy_",)], names=["_dummy_"]
                    )
                    assert _nonone_(
                        df.columns
                    ), f"None in {df.columns.names}, k={k!r}, df={df}"
            assert isinstance(df, pandas.DataFrame), f"Unexpected type {type(df)} for k={k!r}"
            assert _nonone_(df.index), f"None in {df.index.names}, k={k!r}"
            assert _nonone_(df.columns), f"None in {df.columns.names}, k={k!r}, df={df}"
        assert isinstance(df, pandas.DataFrame), f"Unexpected type {type(df)} for k={k!r}"
        assert _nonone_(df.index), f"None in {df.index.names}, k={k!r}"
        assert _nonone_(df.columns), f"None in {df.columns.names}, k={k!r}"
        aggs[f"agg_{k}"] = df

    # check stat is part of the column otherwise the concatenation fails

    set_names = set()
    for df in aggs.values():
        set_names |= set(df.index.names)

    for k, df in aggs.items():
        assert _nonone_(df.index), f"None in {df.index.names}, k={k!r}"
        assert _nonone_(df.columns), f"None in {df.columns.names}, k={k!r}"
        if len(df.index.names) == len(set_names):
            continue
        missing = set_names - set(df.index.names)
        for g in missing:
            df.index = _add_level(df.index, g, k.replace("agg_", ""))

    aggs = {k: _sort_index_level(df, debug=k) for k, df in aggs.items()}

    # concatenation
    dfs = pandas.concat(list(aggs.values()), axis=0)
    assert None not in dfs.index.names, f"None in dfs.index.names={dfs.index.names}"
    assert (
        None not in dfs.columns.names
    ), f"None in dfs.columns.names={dfs.columns.names}, aggs: {set(aggs)}"
    names = list(dfs.columns.names)
    dfs = dfs.unstack(key)
    keep_columns = dfs
    assert None not in dfs.index.names, f"None in dfs.index.names={dfs.index.names}"
    for n in names:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=FutureWarning)
            dfs = dfs.stack(n)
        assert (
            None not in dfs.index.names
        ), f"None in dfs.index.names={dfs.index.names}, n={n!r}"

    assert None not in dfs.index.names, f"None in dfs.index.names={dfs.index.names}"
    dfs = dfs.sort_index(axis=1).sort_index(axis=0)
    keep_columns = keep_columns.sort_index(axis=1).sort_index(axis=0)
    assert None not in dfs.index.names, f"None in dfs.index.names={dfs.index.names}"
    return dfs, keep_columns


def _reverse_column_names_order(
    df: pandas.DataFrame,  # :
    name: Optional[str] = None,
) -> pandas.DataFrame:  # :
    if len(df.columns.names) <= 1:
        return df
    col_names = df.columns.names
    assert isinstance(col_names, list), f"Unexpected type for {df.columns.names!r}"
    return df


def _select_metrics(
    df: pandas.DataFrame,
    select: List[Dict[str, str]],
    prefix: Optional[str] = None,
) -> pandas.DataFrame:
    suites = set(df.columns)
    rows = []
    names = list(df.index.names)
    set_names = set(names)
    for i in df.index.tolist():
        rows.append(set(dict(zip(names, i)).items()))

    subset = [(s, set({k: v for k, v in s.items() if k in set_names}.items())) for s in select]

    keep = []
    for i, row in enumerate(rows):
        for j, (d, s) in enumerate(subset):
            if (s & row) == s:
                keep.append((i, d["new_name"], d["unit"], d.get("order", j), d["help"]))
                break

    dfi = df.iloc[[k[0] for k in keep]].reset_index(drop=False).copy()
    has_suite = "suite" in dfi.columns.names and "exporter" in dfi.columns.names

    def _mk(c):
        if not has_suite:
            return c
        cc = []
        for n in dfi.columns.names:
            if n == "exporter":
                cc.append(c)
            elif c in ("#order", "METRIC"):
                cc.append("")
            else:
                cc.append("~")
        return tuple(cc)

    dfi[_mk("METRIC")] = [k[1] for k in keep]
    dfi[_mk("#order")] = [k[3] for k in keep]
    dfi[_mk("unit")] = [k[2] for k in keep]
    dfi[_mk("~help")] = [k[4] for k in keep]
    dfi = dfi.copy()

    dd_ = set(select[0].keys())
    dd = set()
    for d in dd_:
        cs = [c for c in dfi.columns if d in c]
        if cs:
            dd.add(cs[0])
    cols = [
        *[c for c in dfi.columns if "#order" in c],
        *[c for c in dfi.columns if "METRIC" in c],
    ]
    skip = {*cols, *[c for c in dfi.columns if "unit" in c or "~help" in c]}
    for c in dfi.columns:
        if c in skip or c in dd:
            continue
        cols.append(c)
    cols.extend([c for c in dfi.columns if "unit" in c])
    cols.extend([c for c in dfi.columns if "~help" in c])
    dfi = dfi[cols].sort_values(cols[:-2])
    if prefix == "SIMPLE":
        # flat table
        import pandas

        col_suites = [c for c in dfi.columns if c in suites]
        dfs = []
        for cs in col_suites:
            to_drop = [c for c in col_suites if c != cs]
            if to_drop:
                df = dfi.drop(to_drop, axis=1).copy()
            else:
                df = dfi.copy()
            df["suite"] = cs
            df["value"] = df[cs]
            df = df.drop(cs, axis=1)
            df.columns = [str(c) for c in df.columns]
            dfs.append(df[~df["value"].isna()])
        dfi = pandas.concat(dfs, axis=0).reset_index(drop=True)
    return dfi.sort_index(axis=1), suites


def _filter_data(
    df: pandas.DataFrame,
    filter_in: Optional[str] = None,
    filter_out: Optional[str] = None,
) -> pandas.DataFrame:
    """
    Argument `filter` follows the syntax
    ``<column1>:<fmt1>/<column2>:<fmt2>``.

    The format is the following:

    * a value or a set of values separated by ``;``
    """
    if not filter_in and not filter_out:
        return df

    def _f(fmt):
        cond = {}
        if isinstance(fmt, str):
            cols = fmt.split("/")
            for c in cols:
                assert ":" in c, f"Unexpected value {c!r} in fmt={fmt!r}"
                spl = c.split(":")
                assert len(spl) == 2, f"Unexpected value {c!r} in fmt={fmt!r}"
                name, fil = spl
                cond[name] = FILTERS[fil] if fil in FILTERS else set(fil.split(";"))
        return cond

    if filter_in:
        cond = _f(filter_in)
        assert isinstance(cond, dict), f"Unexpected type {type(cond)} for fmt={filter_in!r}"
        for k, v in cond.items():
            if k not in df.columns:
                continue
            df = df[df[k].isin(v)]

    if filter_out:
        cond = _f(filter_out)
        assert isinstance(cond, dict), f"Unexpected type {type(cond)} for fmt={filter_out!r}"
        for k, v in cond.items():
            if k not in df.columns:
                continue
            df = df[~df[k].isin(v)]
    return df


def _select_model_metrics(
    res: Dict[str, pandas.DataFrame],
    select: List[Dict[str, str]],
    stack_levels: Sequence[str],
) -> pandas.DataFrame:
    import pandas

    concat = []
    for i, metric in enumerate(select):
        cat, stat, new_name, agg = (
            metric["cat"],
            metric["stat"],
            metric["new_name"],
            metric["agg"],
        )
        if new_name.startswith("average "):
            new_name = new_name[len("average ") :]
        if agg in {"TOTAL", "COUNT", "COUNT%", "MAX", "SUM", "NAN"}:
            continue
        name = f"{cat}_{stat}"
        if name not in res:
            continue
        df = res[name].copy()
        cols = list(df.columns)
        if len(cols) == 1:
            col = (cols[0],) if isinstance(cols[0], (str, int)) else tuple(cols[0])
            col = (i, cat, stat, new_name, *col)
            names = ["#order", "cat", "stat", "full_name", *df.columns.names]
            df.columns = pandas.MultiIndex.from_tuples([col], names=names)
            concat.append(df)
        else:
            cols = [((c,) if isinstance(c, str) else tuple(c)) for c in cols]
            cols = [(i, cat, stat, new_name, *c) for c in cols]
            names = ["#order", "cat", "stat", "full_name", *df.columns.names]
            df.columns = pandas.MultiIndex.from_tuples(cols, names=names)
            concat.append(df)
    df = pandas.concat(concat, axis=1)
    if stack_levels:
        for c in stack_levels:
            if c in df.columns.names:
                with warnings.catch_warnings():
                    warnings.simplefilter(
                        "ignore", category=(FutureWarning, PerformanceWarning)
                    )
                    df = df.stack(c, dropna=np.nan)
    df = df.sort_index(axis=1)
    return df


def _compute_correlations(
    df: pandas.DataFrame,
    model_column: List[str],
    exporter_column: List[str],
    columns: List[str],
    verbose: int = 0,
) -> Dict[str, pandas.DataFrame]:
    """
    Computes correlations metrics.

    :param df: raw data
    :param model_column: what defines a model
    :param exporter_column: what defines an exporter
    :param columns: columns to look into
    :param verbose: verbosity
    :return: dictionary of dataframes
    """
    df = df.copy()
    df["_runs_"] = 1
    assert "_runs_" not in [
        *model_column,
        *exporter_column,
        *columns,
    ], f"Name '_runs_' is already taken in {sorted(df.columns)}"

    unique_exporter = df[[*exporter_column, "_runs_"]]
    n_runs = (
        unique_exporter.groupby(exporter_column, as_index=False).sum().reset_index(drop=True)
    )

    res = {"RUNS": n_runs}
    name_i = [f"{c}_i" for c in exporter_column]
    name_j = [f"{c}_j" for c in exporter_column]
    nonans = {}

    piv = None
    for c in columns:
        if verbose:
            print(f"[_compute_correlations] process {c!r}")
        piv = df.pivot(index=model_column, columns=exporter_column, values=[c])

        obs = []
        for i in range(piv.shape[1]):
            for j in range(piv.shape[1]):
                ci = piv.columns[i]
                cj = piv.columns[j]
                ni = ~piv[ci].isna()
                nj = ~piv[cj].isna()
                nonan_ = (ni & nj).apply(lambda x: 1 if x else 0)
                nonan = nonan_.sum()
                sumij = (piv[ci].fillna(0) * nonan_).sum()
                o = dict(
                    zip(
                        [*name_i, *name_j, f"nonan_{c}", f"sum_{c}"],
                        [*ci[1:], *cj[1:], nonan, sumij],
                    )
                )
                if c == "time_latency":
                    # Measuring the best
                    winners = (piv[ci].fillna(0) < (piv[cj].fillna(0) * 0.98)).astype(
                        int
                    ) * nonan_
                    o["win_latency"] = winners.sum()
                elif c == "discrepancies_abs":
                    # Measuring the best
                    winners = (piv[ci].fillna(0) < (piv[cj].fillna(0) * 0.75)).astype(
                        int
                    ) * nonan_
                    o["win_disc_abs"] = winners.sum()
                obs.append(o)
                key = ci[1:], cj[1:]
                if key not in nonans:
                    nonans[key] = nonan_
                else:
                    nonans[key] &= nonan_
        res[f"c_{c}"] = pandas.DataFrame(obs)

    res_join = None
    if piv is not None:
        obs = []
        for i in range(piv.shape[1]):
            for j in range(piv.shape[1]):
                ci = piv.columns[i][1:]
                cj = piv.columns[j][1:]
                o = dict(
                    zip(
                        [*name_i, *name_j, "nonan"],
                        [*ci, *cj, nonans[ci, cj].sum()],
                    )
                )
                obs.append(o)
        res_join = {"nonan": pandas.DataFrame(obs)}

    piv = None
    for c in columns:
        if verbose:
            print(f"[_compute_correlations] process 2 {c!r}")
        piv = df.pivot(index=model_column, columns=exporter_column, values=[c])

        obs = []
        for i in range(piv.shape[1]):
            for j in range(piv.shape[1]):
                ci = piv.columns[i]
                cj = piv.columns[j]
                sumij = (piv[ci].fillna(0) * nonans[ci[1:], cj[1:]]).sum()
                o = dict(
                    zip(
                        [*name_i, *name_j, f"sum_{c}"],
                        [*ci[1:], *cj[1:], sumij],
                    )
                )
                if c == "time_latency":
                    # Measuring the best
                    winners = (piv[ci].fillna(0) < (piv[cj].fillna(0) * 0.98)).astype(
                        int
                    ) * nonans[ci[1:], cj[1:]]
                    o["win_latency"] = winners.sum()
                elif c == "discrepancies_abs":
                    # Measuring the best
                    winners = (piv[ci].fillna(0) < (piv[cj].fillna(0) * 0.75)).astype(
                        int
                    ) * nonans[ci[1:], cj[1:]]
                    o["win_disc_abs"] = winners.sum()
                obs.append(o)
        res_join[f"c_{c}"] = pandas.DataFrame(obs)

    if res_join is not None:
        index_columns = [*name_i, *name_j]
        joined = res_join["nonan"].set_index(index_columns)
        for c in columns:
            if verbose:
                print(f"[_compute_correlations] joins {c!r}")
            joined = joined.join(res_join[f"c_{c}"].set_index(index_columns), how="outer")
        if "win_latency" in joined.columns:
            res["LATENCY"] = pandas.pivot_table(
                joined, index=name_i, columns=name_j, values="win_latency"
            )
        if "win_disc_abs" in joined.columns:
            res["DISC_ABS"] = pandas.pivot_table(
                joined, index=name_i, columns=name_j, values="win_disc_abs"
            )
        res["JOINED"] = joined.reset_index(drop=False)
    return res


def _fix_report_piv(
    piv: pandas.DataFrame, agg: bool = False
) -> Tuple[pandas.DataFrame, pandas.DataFrame]:
    if agg:
        piv = piv[piv.index != (15, "average export time")]
        piv = piv[piv.index != (16, "average speedup (geo)")]

    # simplify dates
    indices = list(enumerate(piv.index))
    dates = [row[0] for row in indices if "date" in row[1]]
    piv.iloc[dates, :] = piv.iloc[dates, :].map(lambda s: s[:10] if isinstance(s, str) else s)

    # add speed by latency
    latencies = [row[0] for row in indices if "total latency time exported model" in row[1]]
    insert_at = []
    for ind in latencies:
        index = piv.index[ind]
        index = (*index[:-1], "speedup weighted by latency")
        speedup = piv.iloc[ind + 1].values / piv.iloc[ind].values
        mindex = pandas.MultiIndex.from_tuples([index], names=piv.index.names)
        add = pandas.DataFrame([speedup.tolist()], columns=piv.columns, index=mindex)
        insert_at.append(add)
    if insert_at:
        piv = pandas.concat([piv, *insert_at], axis=0)
        piv = piv.sort_index()
        return piv, pandas.concat(insert_at, axis=0)
    return piv, None


def _process_formulas(
    df: pandas.DataFrame,
    formulas: List[str],
    column_keys: List[str],
    new_keys: List[str],
    model: List[str],
    set_columns: List[str],
    verbose: int = 0,
) -> Tuple[pandas.DataFrame, List[str]]:
    report_on = []
    for expr in formulas:
        if verbose:
            print(f"[merge_benchmark_reports] process formula={expr}")

        if expr == "memory_delta":
            if (
                "memory_begin" in set_columns
                and "memory_peak" in set_columns
                and "memory_end" in set_columns
            ):
                df["memory_peak_cpu_pp"] = (
                    np.maximum(df["memory_peak"], df["memory_end"]) - df["memory_begin"]
                )
                report_on.append("memory_peak_cpu_pp")
            delta_gpu = None
            m_gpu = None
            for i in range(1024):
                c1 = f"memory_gpu{i}_begin"
                c2 = f"memory_gpu{i}_peak"
                if c1 in set_columns and c2 in set_columns:
                    d = df[c2] - df[c1]
                    if delta_gpu is None:
                        delta_gpu = d
                        m_gpu = d
                    else:
                        delta_gpu += d
                        m_gpu += d
                else:
                    break
            if delta_gpu is not None:
                df["memory_peak_gpu_pp"] = m_gpu
                df["memory_delta_peak_gpu_pp"] = delta_gpu * 2**20
                report_on.append("memory_peak_gpu_pp")
                report_on.append("memory_delta_peak_gpu_pp")
            if "memory_peak_cpu_pp" in df.columns:
                df["memory_delta_peak_cpu_pp"] = df["memory_peak_cpu_pp"] * 2**20
                report_on.append("memory_delta_peak_cpu_pp")
            continue

        if expr == "memory_peak":
            if (
                "mema_gpu_5_after_export" in set_columns
                and "mema_gpu_4_reset" in set_columns
                and "mema_gpu_1_after_loading" in set_columns
                and "mema_gpu_2_after_warmup" in set_columns
                and "mema_gpu_6_before_session" in set_columns
                and "mema_gpu_8_after_export_warmup" in set_columns
            ):
                col_name = "memory_delta_peak_gpu_export"
                df[col_name] = df["mema_gpu_5_after_export"] - df["mema_gpu_4_reset"]
                report_on.append(col_name)

                col_name = "memory_peak_gpu_export"
                df[col_name] = df["mema_gpu_5_after_export"]
                report_on.append(col_name)

                col_name = "memory_delta_peak_gpu_eager_warmup"
                df[col_name] = df["mema_gpu_2_after_warmup"] - df["mema_gpu_0_before_loading"]
                report_on.append(col_name)

                col_name = "memory_peak_gpu_eager_warmup"
                df[col_name] = df["mema_gpu_2_after_warmup"]
                report_on.append(col_name)

                col_name = "memory_delta_peak_gpu_warmup"
                df[col_name] = (
                    df["mema_gpu_8_after_export_warmup"] - df["mema_gpu_6_before_session"]
                )
                report_on.append(col_name)

                col_name = "memory_peak_gpu_warmup"
                df[col_name] = df["mema_gpu_8_after_export_warmup"]
                report_on.append(col_name)
            continue

        if expr == "pass_rate":
            if "discrepancies_abs" in set_columns and "speedup" in set_columns:
                col = (df["discrepancies_abs"] <= 0.1) & (df["speedup"] >= 0.98)
                if "discrepancies_dynamic_abs" in set_columns and "dynamic" in set_columns:
                    col &= (df["dynamic"] == 0) | (
                        (~df["discrepancies_dynamic_abs"].isna())
                        & (df["discrepancies_dynamic_abs"] <= 0.1)
                    )
                df["status_pass_rate"] = col.astype(int)
                df.loc[df["discrepancies_abs"].isna(), "status_pass_rate"] = np.nan
                report_on.append("status_pass_rate")
            continue

        if expr == "correction":
            if "time_latency_eager" in df.columns and "time_latency" in df.columns:
                weights = df["time_latency"].apply(lambda x: np.nan if np.isnan(x) else 1.0)
                df["time_latency_eager_if_exported_run"] = df["time_latency_eager"] * weights
                report_on.append("time_latency_eager_if_exported_run")
            continue

        if expr == "accuracy_rate":
            if "discrepancies_abs" in set_columns:
                col = df["discrepancies_abs"] <= 0.1
                df["status_accuracy_rate"] = col.astype(int)
                df.loc[df["discrepancies_abs"].isna(), "status_accuracy_rate"] = np.nan
                report_on.append("status_accuracy_rate")
            continue

        if expr == "accuracy_dynamic_rate":
            if "discrepancies_dynamic_abs" in set_columns:
                col = df["discrepancies_dynamic_abs"] <= 0.1
                df["status_accuracy_dynamic_rate"] = col.astype(int)
                df.loc[
                    df["discrepancies_dynamic_abs"].isna(), "status_accuracy_dynamic_rate"
                ] = np.nan
                report_on.append("status_accuracy_dynamic_rate")
            continue

        if expr == "date":
            if "date_start" in set_columns:
                df["status_date"] = (
                    pandas.to_datetime(df["date_start"]).astype("int64").astype(float) / 1e9
                )
                set_columns = set(df.columns)
                report_on.append("status_date")
            continue

        if expr == "status":
            if "time_export_unbiased" in set_columns:
                df["status_convert"] = (~df["time_export_unbiased"].isna()).astype(int)
                report_on.append("status_convert")

            if "discrepancies_dynamic_abs" in set_columns:
                df["status_dynamic"] = (
                    (~df["discrepancies_dynamic_abs"].isna())
                    & (df["discrepancies_dynamic_abs"] <= 0.1)
                ).astype(int)
                report_on.append("status_dynamic")

                df["status_convert_ort_dynamic"] = (
                    ~df["discrepancies_dynamic_abs"].isna()
                ).astype(int)
                mets = []
                for th, mt in itertools.product(["1e-1", "1e-2"], ["abs"]):
                    dis = f"discrepancies_dynamic_{mt}"
                    if dis not in df.columns:
                        continue
                    met = f"status_dynerr{mt[3:]}<{th}"
                    mets.append(met)
                    df[met] = (~df[dis].isna() & (df[dis] < float(th))).astype(int)
                set_columns = set(df.columns)
                report_on.extend(["status_convert_ort_dynamic", *mets])

            if "discrepancies_abs" in set_columns:
                df["status_convert_ort"] = (~df["discrepancies_abs"].isna()).astype(int)
                mets = []
                for th, mt in itertools.product(
                    ["1e-1", "1e-2", "1e-3", "1e-4"], ["abs", "abs_0", "abs_1+"]
                ):
                    dis = f"discrepancies_{mt}"
                    if dis not in df.columns:
                        continue
                    met = f"status_err{mt[3:]}<{th}"
                    mets.append(met)
                    df[met] = (~df[dis].isna() & (df[dis] < float(th))).astype(int)
                df["status_lat<=eager+2%"] = (
                    ~df["discrepancies_abs"].isna()
                    & (df["time_latency"] <= df["time_latency_eager"] * 1.02)
                ).astype(int)
                set_columns = set(df.columns)
                report_on.extend(
                    [
                        "status_convert_ort",
                        *mets,
                        "status_lat<=eager+2%",
                    ]
                )

            continue

        if expr == "control_flow":
            if (
                "exporter" in set_columns
                and "time_export_unbiased" in set_columns
                and ({"export", "export-nostrict", "compile"} & set(df.exporter))
                and len(set(df.exporter)) > 1
            ):
                expo = (
                    "export"
                    if ({"export", "export-nostrict"} & set(df.exporter))
                    else "compile"
                )
                keep = [*model, *new_keys, "time_export_unbiased"]
                gr = df[df.exporter == expo][keep].copy()
                gr["status_control_flow"] = gr["time_export_unbiased"].isna().astype(int)
                gr = gr.drop("time_export_unbiased", axis=1)
                if "opt_patterns" in gr.columns and len(set(gr.opt_patterns)) == 1:
                    on = [
                        k for k in keep[:-1] if k not in {"exporter", "opt_patterns", "rtopt"}
                    ]
                else:
                    on = [k for k in keep[:-1] if k != "exporter"]
                joined = pandas.merge(df, gr, left_on=on, right_on=on, how="left")
                assert (
                    df.shape[0] == joined.shape[0]
                ), f"Shape mismatch after join {df.shape} -> {joined.shape}"
                df = joined.copy()
                for c in column_keys:
                    cc = f"{c}_x"
                    if cc in df.columns:
                        df[c] = df[cc]
                drop = [
                    c
                    for c in [
                        *[f"{c}_x" for c in column_keys],
                        *[f"{c}_y" for c in column_keys],
                    ]
                    if c in df.columns
                ]
                df = df.drop(drop, axis=1)
                set_columns = set(df.columns)
                report_on.append("status_control_flow")
            continue

        if expr == "buckets":
            if (
                "exporter" in set_columns
                and "dynamic" in set_columns
                and "opt_patterns" in set_columns
                and "speedup" in set_columns
                and "torch_script" in set(df.exporter)
                and len(set(df.exporter)) > 1
            ):
                # Do the same with the exporter as a baseline.
                keep = [*model, *new_keys, "speedup"]
                gr = df[
                    (df.exporter == "torch_script")
                    & (df.opt_patterns.isin({"", "-", "none"}))
                    & (df.rtopt == 1)
                ][keep].copy()
                gr = gr[~gr["speedup"].isna()]

                if gr.shape[0] == 0:
                    # No figures for torch_script
                    if verbose:
                        print(
                            f"[merge_benchmark_reports] gr.shape={gr.shape}, "
                            f"unable to compute speedup_script, "
                            f"exporters={set(df.exporter)}, "
                            f"opt_patterns={set(df.opt_patterns)}, "
                            f"dynamic={set(df.dynamic)}, rtopt={set(df.rtopt)}"
                        )
                else:
                    if verbose:
                        print(
                            f"[merge_benchmark_reports] compute speedup_script: "
                            f"gr.shape={gr.shape}"
                        )
                    gr["speedup_script"] = gr["speedup"]
                    gr = gr.drop("speedup", axis=1)

                    on = [
                        k for k in keep[:-1] if k not in {"exporter", "opt_patterns", "rtopt"}
                    ]
                    joined = pandas.merge(df, gr, left_on=on, right_on=on, how="left")

                    assert df.shape[0] == joined.shape[0], (
                        f"Shape mismatch after join {df.shape} -> {joined.shape}, "
                        f"gr.shape={gr.shape}, on={on}. This usually means you have "
                        f"duplicates. You should use keep_more_recent=True"
                    )
                    df = joined.copy()
                    # We cannot replace NaN by -inf here,
                    # that artificially increases the first bucket.
                    df["speedup_increase_script"] = df["speedup"] / df["speedup_script"] - 1
                    report_on.extend(["speedup_script", "speedup_increase_script"])
                    for c in column_keys:
                        cc = f"{c}_x"
                        if cc in df.columns:
                            df[c] = df[cc]
                    drop = [
                        c
                        for c in [
                            *[f"{c}_x" for c in column_keys],
                            *[f"{c}_y" for c in column_keys],
                        ]
                        if c in df.columns
                    ]
                    df = df.drop(drop, axis=1)
                    set_columns = set(df.columns)
                    df["status_lat<=script+2%"] = (
                        df["speedup_increase_script"] >= (1 / 1.02 - 1)
                    ).astype(int)
                    report_on.append("status_lat<=script+2%")

            for c in ["speedup_increase", "speedup_increase_script"]:
                if c not in set_columns:
                    continue
                scale = BUCKET_SCALES
                ind = df["speedup_increase"].isna()
                for i in range(1, len(scale)):
                    val = (df[c] >= scale[i - 1] / 100) & (df[c] < scale[i] / 100)
                    v1 = f"{scale[i-1]}%" if not np.isinf(scale[i - 1]) else ""
                    v2 = f"{scale[i]}%" if not np.isinf(scale[i]) else ""
                    suf = f"[{v1},{v2}[" if v1 and v2 else (f"<{v2}" if v2 else f">={v1}")
                    if c == "speedup_increase_script":
                        d = f"bucket_script {suf}"
                    else:
                        d = f"bucket_{suf}"
                    df[d] = val.astype(int)
                    df.loc[ind, d] = np.nan
                    report_on.append(d)

            # for inductor
            if (
                "exporter" in set_columns
                and "dynamic" in set_columns
                and "opt_patterns" in set_columns
                and "speedup" in set_columns
                and "inductor" in set(df.exporter)
                and len(set(df.exporter)) > 1
            ):
                # Do the same with the exporter as a baseline.
                keep = [*model, *new_keys, "speedup"]
                gr = df[
                    (df.exporter == "inductor")
                    & (df.opt_patterns.isin({"", "-", "none"}))
                    & (df.rtopt == 1)
                ][keep].copy()
                gr = gr[~gr["speedup"].isna()]

                if gr.shape[0] == 0:
                    # No figures for inductor
                    if verbose:
                        print(
                            f"[merge_benchmark_reports] gr.shape={gr.shape}, "
                            f"unable to compute speedup_inductor, "
                            f"exporters={set(df.exporter)}, "
                            f"opt_patterns={set(df.opt_patterns)}, "
                            f"dynamic={set(df.dynamic)}, rtopt={set(df.rtopt)}"
                        )
                else:
                    if verbose:
                        print(
                            f"[merge_benchmark_reports] compute speedup_inductor: "
                            f"gr.shape={gr.shape}"
                        )
                    gr["speedup_inductor"] = gr["speedup"]
                    gr = gr.drop("speedup", axis=1)

                    on = [
                        k for k in keep[:-1] if k not in {"exporter", "opt_patterns", "rtopt"}
                    ]
                    joined = pandas.merge(df, gr, left_on=on, right_on=on, how="left")

                    assert df.shape[0] == joined.shape[0], (
                        f"Shape mismatch after join {df.shape} -> {joined.shape}, "
                        f"gr.shape={gr.shape}, on={on}"
                    )
                    df = joined.copy()
                    df["speedup_increase_inductor"] = (
                        df["speedup"] / df["speedup_inductor"] - 1
                    ).fillna(-np.inf)
                    report_on.extend(["speedup_inductor", "speedup_increase_inductor"])
                    for c in column_keys:
                        cc = f"{c}_x"
                        if cc in df.columns:
                            df[c] = df[cc]
                    drop = [
                        c
                        for c in [
                            *[f"{c}_x" for c in column_keys],
                            *[f"{c}_y" for c in column_keys],
                        ]
                        if c in df.columns
                    ]
                    df = df.drop(drop, axis=1)
                    set_columns = set(df.columns)
                    df["status_lat<=inductor+2%"] = (
                        df["speedup_increase_inductor"] >= (1 / 1.02 - 1)
                    ).astype(int)
                    report_on.append("status_lat<=inductor+2%")

            continue

        if expr == "error":

            def _filter(v):
                lines = v.split("\n")
                new_lines = []
                for line in lines:
                    if len(line) < 5:
                        continue
                    if (
                        "Starting from v4.46 the `logits` model output "
                        "will have the same type as the model"
                    ) in line:
                        continue
                    if "Could not find a CPU kernel and hence can't constant fold" in line:
                        continue
                    if "UserWarning: " in line:
                        continue
                    new_lines.append(line)
                return "\n".join(new_lines)

            def _text_rows(row, err_cols, disc_name):
                if disc_name and not np.isnan(row[disc_name]):
                    return ""
                for c in [
                    "ERR_timeout",
                    "ERR_load",
                    "ERR_warmup_eager",
                    "ERR_export",
                    "ERR_feeds",
                    "ERR_ort",
                    "ERR_warmup",
                    "ERR_std",
                ]:
                    if c not in err_cols:
                        continue
                    v = row[c]
                    if not isinstance(v, str):
                        continue
                    if len(v) > 20:
                        v = _filter(v)
                        if v:
                            return f"{c}: {v}"
                return ""

            add = {}
            err_cols = []
            for c in df.columns:
                if c.startswith("ERR_") and df[c].dtype in (str, object):
                    setup = df[c].str.contains(
                        "Cannot install -r requirements.txt", regex=False
                    )
                    if True in set(setup):
                        add[f"ERR_SETUP_{c[4:]}"] = setup.fillna(0.0).astype(int)
                        report_on.append(f"ERR_SETUP_{c[4:]}")
                    oom = df[c].str.contains("CUDA out of memory", regex=False)
                    if True in set(oom):
                        add[f"ERR_OOM_{c[4:]}"] = oom.fillna(0.0).astype(int)
                        report_on.append(f"ERR_OOM_{c[4:]}")
                    oomort = df[c].str.contains(
                        "onnxruntime::BFCArena::AllocateRawInternal(size_t bool "
                        "onnxruntime::Stream* bool onnxruntime::WaitNotificationFn) "
                        "Failed to allocate memory for requested buffer of size",
                        regex=False,
                    )
                    if True in set(oomort):
                        add[f"ERR_OOMORT_{c[4:]}"] = oomort.fillna(0.0).astype(int)
                        report_on.append(f"ERR_OOMORT_{c[4:]}")
                    acc = df[c].str.contains("Cannot access gated repo for url", regex=False)
                    if True in set(acc):
                        add[f"ERR_HTTP_{c[4:]}"] = acc.fillna(0.0).astype(int)
                        report_on.append(f"ERR_HTTP_{c[4:]}")
                    mem = df[c].str.contains(
                        "Memcpy nodes are added to the graph main_graph "
                        "for CUDAExecutionProvider",
                        regex=False,
                    )
                    if True in set(mem):
                        add[f"ERR_ORTMEMCPY_{c[4:]}"] = mem.fillna(0.0).astype(int)
                        report_on.append(f"ERR_ORTMEMCPY_{c[4:]}")
                    err_cols.append(c)
            if err_cols:
                set_cols = set(err_cols)
                disc_name = "discrepancies_abs"
                if disc_name in df.columns:
                    err_cols.append(disc_name)
                else:
                    disc_name = None
                df["ERR_FIRST"] = df[err_cols].apply(
                    lambda row, set_cols=set_cols, disc_name=disc_name: _text_rows(
                        row, err_cols=set_cols, disc_name=disc_name
                    ),
                    axis=1,
                )
                report_on.append("ERR_FIRST")
            for k, v in add.items():
                df[k] = v
                report_on.append(k)
            continue

        if expr == "export":
            # guess the export time, for some exporter it is the first iteration.
            def unbiased_export(row):
                exporter = row["exporter"]
                time_export_success = row["time_export_success"]
                time_warmup_first_iteration = row["time_warmup_first_iteration"]
                if (
                    time_export_success is None
                    or np.isnan(time_export_success)
                    or time_warmup_first_iteration is None
                    or np.isnan(time_warmup_first_iteration)
                ):
                    return np.nan
                if exporter in {
                    "inductor",
                    "eager",
                    "export",
                    "export-nostrict",
                    "compile",
                    "dort",
                    "cort",
                }:
                    return time_export_success + time_warmup_first_iteration
                return time_export_success

            if (
                "exporter" in set_columns
                and "time_export_success" in set_columns
                and "time_warmup_first_iteration" in set_columns
            ):
                df["time_export_unbiased"] = df.apply(unbiased_export, axis=1)
                report_on.append("time_export_unbiased")
            elif "exporter" in set_columns and "time_export_success" in set_columns:
                # old data
                df["time_export_unbiased"] = df["time_export_success"]
                report_on.append("time_export_unbiased")

            continue

        raise AssertionError(f"Unknown formula {expr!r}")
    return df, report_on


def build_historical_report(
    output: str,
    input_files: List[str],
    verbose: int = 0,
    filter_in: Optional[Any] = None,
    filter_out: Optional[Any] = None,
    pages: Optional[Union[str, List[str]]] = None,
):
    """
    Builds historical graph using the aggregated data (export_simple options).

    :param output: output, an excel file
    :param input_files: input_files
    :param verbose: verbosity
    :param filter_in: filter in some data to make the report smaller (see below)
    :param filter_out: filter out some data to make the report smaller (see below)
    :param pages: list of pages to produce, None for all, if can be a string,
        comma separated values or a of list of strings

    Argument `filter_in` or `filter_out` follows the syntax
    ``<column1>:<fmt1>/<column2>:<fmt2>``.

    The format is the following:

    * a value or a set of values separated by ``;``
    """
    expected_columns = ["METRIC", "suite", "value", "DATE"]
    dfs = []
    for name in input_files:
        if verbose:
            print(f"[build_historical_report] read {name!r}")
        df = pandas.read_csv(name)
        if verbose > 2:
            print(df.head())
        assert all(
            c in df.columns for c in expected_columns
        ), f"Unexpected columns {df.columns} in {name!r}"
        dfs.append(df)

    df = pandas.concat(dfs, axis=0)

    if filter_in or filter_out:
        if verbose:
            print("[merge_benchmark_reports] filtering data")

        df = _filter_data(df, filter_in=filter_in, filter_out=filter_out)

        if verbose:
            print(f"[merge_benchmark_reports] done, new shape={df.shape}")
        if df.shape[0] == 0:
            return {}

    df = df[df.METRIC != "date"]
    df["value"] = df["value"].astype(float)
    df["dtype"] = df["dtype"].fillna("all")
    df = df[~df["value"].isna()]
    exporter = [c for c in ["exporter", "opt_patterns", "dynamic", "dtype"] if c in df.columns]
    if verbose:
        print(f"[build_historical_report] shape={df.shape}, exporter={exporter}")
        print(f"[build_historical_report] unique exporter={set(df.exporter)}")
        print(f"[build_historical_report] suite={set(df.suite)}")

    graphs = {
        "model number": ["number of models", "number of models eager mode"],
        "export number": [
            "export number",
            "number of running models",
            "accuracy number",
            "pass number",
        ],
        "faster number": [
            "number of models equal or faster than eager",
            "number of models equal or faster than inductor",
        ],
        "torch.export.export numbers": [
            "number of models",
            "number of models eager mode",
            "number of failures for torch.export.export",
        ],
        "benchmark time": ["total export time", "benchmark duration"],
        "export time": ["total export time"],
        "speedup": ["speedup weighted by latency", "average speedup (geo)"],
        "discrepancies": ["discrepancies < 0.1", "discrepancies < 0.01"],
        "discrepancies_dynamic": [
            "discrepancies_dynamic < 0.1",
            "discrepancies_dynamic < 0.01",
        ],
        "memory": [
            "average GPU delta peak (export) (torch)",
            "average GPU delta peak (export) (nvidia-smi)",
        ],
    }
    if pages:
        if isinstance(pages, str):
            pages = pages.split(",")
        pages = set(pages)
        graphs = {k: v for k, v in graphs.items() if k in pages}

    if verbose:
        print(f"[build_historical_report] create {output!r}, verbose={verbose}")
    with pandas.ExcelWriter(output, engine="xlsxwriter") as writer:
        export_export = {}
        for k, v in graphs.items():
            if verbose:
                print(f"[build_historical_report] create graph {k!r}, exporter={exporter}")
            sdf = df[df.METRIC.isin(v)]
            if sdf.shape[0] == 0:
                if verbose:
                    print(
                        f"[build_historical_report] empty graph for {k!r}, exporter={exporter}"
                    )
                continue
            sdf = sdf.sort_values([*exporter, "suite", "METRIC", "DATE"])
            sdf = sdf[[*exporter, "suite", "METRIC", "DATE", "value"]].copy()
            if verbose:
                print(
                    f"[build_historical_report] shape={sdf.shape}, metrics={set(sdf.METRIC)}"
                )
                if verbose > 2:
                    print(sdf.head())
                    print(sdf.tail())

            try:
                piv = sdf.pivot(
                    index=[*exporter, "suite", "DATE"], columns="METRIC", values="value"
                )
            except ValueError as e:
                cc = [*exporter, "suite", "DATE"]
                dg = sdf.copy()
                dg["__C__"] = 1
                under = dg.groupby(cc).count()[["__C__"]]
                under = under[under["__C__"] > 1]
                raise ValueError(f"Ambiguities for columns {cc}\n{under}") from e

            subset = piv.reset_index(drop=False)
            subset.to_excel(writer, sheet_name=k, index=False)
            idate = list(subset.columns).index("DATE")
            if verbose > 2:
                print(f"[build_historical_report] shape={subset.shape}")
                print(subset.head())
                print(subset.tail())

            workbook = writer.book
            worksheet = writer.sheets[k]

            locations_cols = {}
            locations_rows = {}

            i = 0
            while i < subset.shape[0]:
                suite = subset.loc[i, "suite"]
                ex = subset.loc[i, "exporter"]
                optim = subset.loc[i, "opt_patterns"] if "opt_patterns" in exporter else ""
                dynamic = subset.loc[i, "dynamic"] if "dynamic" in exporter else ""
                if suite not in locations_rows:
                    locations_rows[suite] = (
                        (max(locations_rows.values()) + 15) if locations_rows else 0
                    )
                key = int(dynamic), ex, optim
                if key not in locations_cols:
                    locations_cols[key] = (
                        (max(locations_cols.values()) + 8) if locations_cols else 0
                    )

                j = i + 1
                while j < subset.shape[0] and subset.loc[j, "suite"] == suite:
                    j += 1

                if verbose:
                    print(
                        f"[build_historical_report] + {suite},{ex},{optim},{dynamic}, "
                        f"d={idate}, add chart from row {i} to {j} ({k!r})"
                    )

                chart = workbook.add_chart({"type": "line"})
                chart2 = workbook.add_chart({"type": "line"})
                for col in v:
                    if col not in subset.columns:
                        continue
                    ivalue = list(subset.columns).index(col)
                    if verbose:
                        print(f"[build_historical_report] ++ serie {col!r}")
                    kwargs = {
                        "name": [k, 0, ivalue],
                        "categories": [k, i + 1, idate, j, idate],
                        "values": [k, i + 1, ivalue, j, ivalue],
                        "line": {"width": 2},
                        "marker": {"type": "diamond"},
                    }
                    chart.add_series(kwargs)
                    chart2.add_series(kwargs)

                chart.set_x_axis({"name": "date", "date_axis": True})
                chart.set_y_axis({"name": k, "major_gridlines": {"visible": False}})
                chart.set_legend({"position": "top"})
                title = f"{suite} - {ex} +{optim}{' dynamic shape' if dynamic else ''}"
                chart.set_title({"name": title})
                chart2.set_title({"name": f"{suite} - {k}"})
                x, y = locations_cols[key], locations_rows[suite]
                place = (
                    f"{chr(65 + x)}{y + 1}"
                    if x < 26
                    else f"{chr(64 + x // 26)}{chr(65 + (x % 26))}{y + 1}"
                )
                if verbose:
                    print(
                        f"[build_historical_report] insert at "
                        f"{place} add title {k}: {title!r}"
                    )
                worksheet.insert_chart(place, chart)
                export_export[key, suite, k] = (chart2, f"{suite} - {k}")
                i = j

        # second round
        locations_cols = set()
        locations_rows = set()
        pages = set()
        for k in export_export:
            key, suite, kind = k
            skey = "-".join(map(str, key))
            pages.add(skey)
            locations_cols.add(suite)
            locations_rows.add(kind)

        pages = sorted(pages)
        locations_rows = sorted(locations_rows)
        locations_cols = sorted(locations_cols)
        for p in pages:
            pandas.DataFrame({p: list(range(10))}).to_excel(writer, sheet_name=p, index=False)
        workbook = writer.book
        for k, (chart, title) in export_export.items():
            key, suite, kind = k
            skey = "-".join(map(str, key))
            worksheet = writer.sheets[skey]
            x = locations_cols.index(suite) * 8
            y = locations_rows.index(kind) * 15
            place = (
                f"{chr(65 + x)}{y + 1}"
                if x < 26
                else f"{chr(64 + x // 26)}{chr(65 + (x % 26))}{y + 1}"
            )
            if verbose:
                print(f"[build_historical_report] insert on {skey} at {place}: {title}")
            worksheet.insert_chart(place, chart)
