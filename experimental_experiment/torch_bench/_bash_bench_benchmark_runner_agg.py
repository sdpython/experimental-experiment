import glob
import itertools
import pprint
import time
import warnings
from collections import Counter
from typing import Dict, List, Optional, Sequence, Set, Union

import numpy as np
import pandas
from pandas.errors import PerformanceWarning

BUCKET_SCALES = [-np.inf, -20, -10, -5, -2, 0, 2, 5, 10, 20, np.inf]
BUCKETS = [
    "<-20%",
    "[-20%,-10%[",
    "[-10%,-5%[",
    "[-5%,-2%[",
    "[-2%,0%[",
    "[0%,2%[",
    "[2%,5%[",
    "[5%,10%[",
    "[10%,20%[",
    ">=20%",
    "script <-20%",
    "script [-20%,-10%[",
    "script [-10%,-5%[",
    "script [-5%,-2%[",
    "script [-2%,0%[",
    "script [0%,2%[",
    "script [2%,5%[",
    "script [5%,10%[",
    "script [10%,20%[",
    "script >=20%",
]


def _SELECTED_FEATURES():
    features = [
        dict(
            cat="status",
            stat="date",
            agg="MAX",
            new_name="date",
            unit="date",
            help="Most recent date involved",
            simple=True,
        ),
        dict(
            cat="time",
            stat="ITER",
            agg="SUM",
            new_name="benchmark duration",
            unit="s",
            help="Total duration of the benchmark",
            simple=True,
        ),
        dict(
            cat="time",
            stat="ITER",
            agg="TOTAL",
            new_name="number of models",
            unit="N",
            help="Number of models evaluated in this document.",
            simple=True,
        ),
        dict(
            cat="status",
            stat="control_flow",
            agg="SUM",
            new_name="number of control flow",
            unit="N",
            help="torch.export.export does not work because of a "
            "control flow, in practice, this column means torch.export.export "
            "succeeds, this metric is only available if the data of exporter "
            "'export' or 'compile' is aggregated",
            simple=True,
        ),
        dict(
            cat="time",
            agg="COUNT",
            stat="export_success",
            new_name="export number",
            unit="N",
            help="Number of models successfully converted into ONNX. "
            "The ONNX model may not be run through onnxruntime or with "
            "significant discrepancies.",
            simple=True,
        ),
        dict(
            cat="speedup",
            agg="COUNT",
            stat="increase",
            new_name="number of running models",
            unit="N",
            help="Number of models converted and running with onnxruntime. "
            "The outputs may be right or wrong. Unit test ensures every aten functions "
            "is correctly converted but the combination may produce outputs "
            "with higher discrepancies than expected.",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="accuracy_rate",
            new_name="accuracy number",
            unit="N",
            help="Number of models successfully converted into ONNX. "
            "It may be slow but the discrepancies are < 0.1.",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="pass_rate",
            new_name="pass number",
            unit="N",
            help="Number of models successfully converted into ONNX, "
            "with a maximum discrepancy < 0.1 and a speedup > 0.98.",
            simple=True,
        ),
        dict(
            cat="time",
            agg="SUM",
            stat="export_success",
            new_name="total export time",
            unit="s",
            help="Total export time when the export succeeds. "
            "The model may not run through onnxruntime and the model "
            "may produce higher discrepancies than expected (lower is better).",
            simple=True,
        ),
        dict(
            cat="time",
            agg="SUM",
            stat="latency",
            new_name="total time ORT",
            unit="x",
            help="Total latency time with onnxruntime",
            simple=True,
        ),
        dict(
            cat="time",
            agg="SUM",
            stat="latency_eager_if_ort",
            new_name="total time eager / ORT",
            unit="x",
            help="Total latency of eager mode knowing that onnxruntime is running",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="lat<=script+2%",
            new_name="number of models equal or faster than torch.script",
            unit="N",
            help="Number of models successfully converted with torch.script "
            "and the other exporter, and the second exporter is as fast or faster "
            "than torch.script.",
            simple=True,
        ),
        dict(
            cat="status",
            agg="SUM",
            stat="lat<=eager+2%",
            new_name="number of models equal or faster than eager",
            unit="N",
            help="Number of models as fast or faster than torch eager mode.",
            simple=True,
        ),
        # average
        dict(
            cat="time",
            agg="MEAN",
            stat="export_success",
            new_name="average export time",
            unit="s",
            help="Average export time when the export succeeds. "
            "The model may not run through onnxruntime and the model "
            "may produce higher discrepancies than expected (lower is better).",
            simple=True,
        ),
        dict(
            cat="speedup",
            agg="GEO-MEAN",
            stat="1speedup",
            new_name="average speedup (geo)",
            unit="x",
            help="Geometric mean of all speedup for all model converted and runnning.",
            simple=True,
        ),
        # e-1
        dict(
            cat="status",
            agg="SUM",
            stat="err<1e-1",
            new_name="discrepancies < 0.1",
            unit="N",
            help="Number of models for which the maximum discrepancies is "
            "below 0.1 for all outputs.",
            simple=True,
        ),
        # e-2
        dict(
            cat="status",
            agg="SUM",
            stat="err<1e-2",
            new_name="discrepancies < 0.01",
            unit="N",
            help="Number of models for which the maximum discrepancies is "
            "below 0.01 for all outputs.",
            simple=True,
        ),
        dict(
            cat="time",
            stat="ITER",
            agg="MEAN",
            new_name="average iteration time",
            unit="s",
            help="Average total time per model and scenario. "
            "It usually reflects how long the export time is (lower is better).",
        ),
        dict(
            cat="discrepancies",
            stat="avg",
            agg="MEAN",
            new_name="average average discrepancies",
            unit="f",
            help="Average of average absolute discrepancies "
            "assuming it can be measured (lower is better).",
        ),
        dict(
            cat="discrepancies",
            stat="abs",
            agg="MEAN",
            new_name="average absolute discrepancies",
            unit="f",
            help="Average maximum absolute discrepancies "
            "assuming it can be measured (lower is better).",
        ),
        dict(
            cat="time",
            agg="MEAN",
            stat="latency_eager",
            new_name="average latency eager",
            unit="s",
            help="Average latency for eager mode (lower is better)",
        ),
        dict(
            cat="time",
            agg="MEAN",
            stat="latency",
            new_name="average latency ort",
            unit="s",
            help="Average latency for onnxruntime (lower is better)",
        ),
        dict(
            cat="onnx",
            agg="MEAN",
            stat="weight_size_torch",
            new_name="average weight size",
            unit="bytes",
            help="Average parameters size, this gives a kind of order "
            "of magnitude for the memory peak",
        ),
        dict(
            cat="onnx",
            agg="MEAN",
            stat="weight_size_proto",
            new_name="average weight size in ModelProto",
            unit="bytes",
            help="Average parameters size in the model proto, "
            "this gives a kind of order of magnitude for the memory peak "
            "this should be close to the parameter size",
        ),
        dict(
            cat="onnx",
            agg="MAX",
            stat="weight_size_torch",
            new_name="maximum weight size",
            unit="bytes",
            help="Maximum parameters size, "
            "useful to guess how much this machine can handle",
        ),
        dict(
            cat="onnx",
            agg="MAX",
            stat="weight_size_proto",
            new_name="maximum weight size in ModelProto",
            unit="bytes",
            help="Maximum parameters size in the model proto, "
            "useful to guess how much this machine can handle",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="peak_gpu_eager_warmup",
            new_name="average GPU peak (eager warmup)",
            unit="bytes",
            help="Average GPU peak while warming up eager mode (torch metric)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_gpu_eager_warmup",
            new_name="average GPU delta peak (eager warmup)",
            unit="bytes",
            help="Average GPU peak of new allocated memory while warming up eager "
            "mode (torch metric)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="peak_gpu_warmup",
            new_name="average GPU peak (warmup)",
            unit="bytes",
            help="Average GPU peak while warming up onnxruntime (torch metric)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_gpu_warmup",
            new_name="average GPU delta peak (warmup)",
            unit="bytes",
            help="Average GPU peak of new allocated memory while warming up "
            "onnxruntime (torch metric)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_cpu_pp",
            new_name="average CPU delta peak (export)",
            unit="bytes",
            help="Average CPU peak of new allocated memory while converting "
            "the model (measured in a secondary process)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_gpu_export",
            new_name="average GPU delta peak (export) (torch)",
            unit="bytes",
            help="Average GPU peak of new allocated memory while "
            "converting the model (torch metric)",
            simple=True,
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_gpu_pp",
            new_name="average GPU delta peak (export) (nvidia-smi)",
            unit="bytes",
            help="Average GPU peak of new allocated memory while "
            "converting the model (measured in a secondary process)",
            simple=True,
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="peak_gpu_export",
            new_name="average GPU peak (export)",
            unit="bytes",
            help="Average GPU peak while converting the model (torch metric)",
        ),
        dict(
            cat="memory",
            agg="MEAN",
            stat="delta_peak_gpu_export",
            new_name="average delta GPU peak (export)",
            unit="bytes",
            help="Average GPU peak of new allocated memory "
            "while converting the model (torch metric)",
            simple=True,
        ),
        dict(
            cat="speedup",
            agg="MEAN",
            stat="increase",
            new_name="average speedup increase",
            unit="%",
            help="Average speedup increase compare to eager mode.",
        ),
    ]
    for b in BUCKETS:
        s = "script" in b
        bs = b.replace("script ", "")
        ag = "torch_script" if s else "eager"
        features.append(
            dict(
                cat="bucket",
                agg="SUM",
                stat=b,
                new_name=f"speedup/script in {bs}" if s else f"speedup in {bs}",
                unit="N",
                help=f"Number of models whose speedup against {ag} "
                f"falls into this interval",
                simple=True,
            )
        )
    return features


SELECTED_FEATURES = _SELECTED_FEATURES()

FILTERS = {
    "HG": [
        "AlbertForMaskedLM",
        "AlbertForQuestionAnswering",
        "AllenaiLongformerBase",
        "BartForCausalLM",
        "BartForConditionalGeneration",
        "BertForMaskedLM",
        "BertForQuestionAnswering",
        "BlenderbotForCausalLM",
        "BlenderbotSmallForCausalLM",
        "BlenderbotSmallForConditionalGeneration",
        "CamemBert",
        "DebertaForMaskedLM",
        "DebertaForQuestionAnswering",
        "DebertaV2ForMaskedLM",
        "DebertaV2ForQuestionAnswering",
        "DistilBertForMaskedLM",
        "DistilBertForQuestionAnswering",
        "DistillGPT2",
        "ElectraForCausalLM",
        "ElectraForQuestionAnswering",
        "GPT2ForSequenceClassification",
        "GoogleFnet",
        "LayoutLMForMaskedLM",
        "LayoutLMForSequenceClassification",
        "M2M100ForConditionalGeneration",
        "MBartForCausalLM",
        "MBartForConditionalGeneration",
        "MT5ForConditionalGeneration",
        "MegatronBertForCausalLM",
        "MegatronBertForQuestionAnswering",
        "MobileBertForMaskedLM",
        "MobileBertForQuestionAnswering",
        "OPTForCausalLM",
        "PLBartForCausalLM",
        "PLBartForConditionalGeneration",
        "PegasusForCausalLM",
        "PegasusForConditionalGeneration",
        "RobertaForCausalLM",
        "RobertaForQuestionAnswering",
        "Speech2Text2ForCausalLM",
        "T5ForConditionalGeneration",
        "T5Small",
        "TrOCRForCausalLM",
        "XGLMForCausalLM",
        "XLNetLMHeadModel",
        "YituTechConvBert",
    ]
}


def _nonone_(index):
    return None not in index.names


def _key(v):
    if isinstance(v, (int, float)):
        return v, ""
    if isinstance(v, str):
        return 1e10, v
    if isinstance(v, tuple):
        return (1e10, *v)
    raise AssertionError(f"Unexpected type for v={v!r}, type is {type(v)}")


def sort_index_key(index):
    return pandas.Index(_key(v) for v in index)


def _isnan(x):
    if x is None:
        return True
    if isinstance(x, str):
        return x == ""
    try:
        return np.isnan(x)
    except TypeError:
        return False


def _isinf(x):
    if x is None:
        return True
    if isinstance(x, str):
        return x == "inf"
    try:
        return np.isinf(x)
    except TypeError:
        return False


def _column_name(ci: int) -> str:
    if ci < 26:
        return chr(65 + ci)
    a = ci // 26
    b = ci % 26
    return f"{chr(65+a)}{chr(65+b)}"


def _apply_excel_style(
    res: Dict[str, pandas.DataFrame],
    writer: pandas.ExcelWriter,
    verbose: int = 0,
):
    from openpyxl.styles import Alignment, Font, PatternFill, numbers

    bold_font = Font(bold=True)
    alignment = Alignment(horizontal="left")
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    red = Font(color="FF0000")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    redf = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    lasts = {}
    for k, v in res.items():
        if verbose:
            print(f"[_apply_excel_style] apply style on {k!r}, shape={v.shape}")
        sheet = writer.sheets[k]
        if 0 in v.shape:
            continue

        if k in ("0main", "0main_base"):
            for c in "AB":
                sheet.column_dimensions[c].width = 40
                sheet.column_dimensions[c].alignment = alignment
            for cell in sheet[1]:
                cell.font = bold_font
                cell.alignment = alignment
            continue

        if k in {"0raw", "AGG", "AGG2", "0raw_base"}:
            continue

        n_cols = (
            1 if isinstance(v.index[0], (str, int, np.int64, np.int32)) else len(v.index[0])
        )
        n_rows = (
            1
            if isinstance(v.columns[0], (str, int, np.int64, np.int32))
            else len(v.columns[0])
        )

        for i in range(n_cols):
            sheet.column_dimensions[_column_name(i)].width = 40

        first_row = None
        first_col = None
        if v.index.names == [None]:
            first_row = len(v.columns.names)
            first_col = len(v.columns.names)
            if verbose > 1:
                print(
                    f"[_apply_excel_style] k={k!r}, first={first_row},{first_col}, "
                    f"columns.names={v.columns.names}"
                )
        else:
            look = v.iloc[0, 0]

            values = []
            for row in sheet.iter_rows(
                min_row=1,
                max_row=n_rows + n_cols + 1,
                min_col=1,
                max_col=n_rows + n_cols + 1,
            ):
                for cell in row:
                    if hasattr(cell, "col_idx") and (
                        cell.value == look
                        or (_isnan(cell.value) and _isnan(look))
                        or (_isinf(cell.value) and _isinf(look))
                    ):
                        first_row = cell.row
                        first_col = cell.col_idx if hasattr(cell, "col_idx") else first_col
                        break
                    values.append(cell.value)
                if first_row is not None:
                    break

            if verbose > 1:
                print(
                    f"[_apply_excel_style] k={k!r}, first={first_row},{first_col}, "
                    f"index.names={v.index.names}, columns.names={v.columns.names}"
                )

            assert first_row is not None and first_col is not None, (
                f"Unable to find the first value in {k!r}, first_row={first_row}, "
                f"first_col={first_col}, look={look!r} ({type(look)}), "
                f"n_rows={n_rows}, n_cols={n_cols}, values={values}, "
                f"iloc[:3,:3]={v.iloc[:3, :3]}, v.index.names={v.index.names}, "
                f"v.columns={v.columns}"
            )

        last_row = first_row + v.shape[0] + 1
        last_col = first_col + v.shape[1] + 1
        lasts[k] = (last_row, last_col)
        for row in sheet.iter_rows(
            min_row=first_row,
            max_row=last_row,
            min_col=first_col,
            max_col=last_col,
        ):
            for cell in row:
                cell.alignment = alignment
                cell.font = bold_font
        for i in range(n_rows):
            for cell in sheet[i + 1]:
                cell.font = bold_font

        if k == "memory":
            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    cell.number_format = numbers.FORMAT_NUMBER
            continue

        if k == "ERR":
            for i in range(n_cols + v.shape[1]):
                sheet.column_dimensions[_column_name(i)].alignment = alignment
                sheet.column_dimensions[_column_name(i)].width = 50
                if i >= 25:
                    break
            continue

        if k in ("TIME_ITER", "discrepancies", "speedup"):
            c1, c2 = None, None
            qpb, qpt, fmtp = None, None, None
            if k == "TIME_ITER":
                qt = np.quantile(v.values, 0.9)
                qb = None
                fmt = numbers.FORMAT_NUMBER_00
            elif k == "speedup":
                debug_values = []
                for row in sheet.iter_rows(
                    min_row=0,
                    max_row=30,
                    min_col=first_col,
                    max_col=last_col,
                ):
                    found = None
                    for cell in row:
                        debug_values.append(cell.value)
                        if cell.value == "increase":
                            c1 = cell.col_idx if c1 is None else min(cell.col_idx, c1)
                            c2 = cell.col_idx if c2 is None else max(cell.col_idx, c2)
                            found = cell
                            last_idx = cell.col_idx
                        elif cell.value is None:
                            # merged cell
                            if found is not None:
                                last_idx += 1
                                cid = getattr(cell, "col_idx", last_idx)
                                c2 = max(cid, c2)
                        else:
                            found = None

                assert (
                    c1 is not None and c2 is not None and c1 <= c2
                ), f"Unexpected value for c1={c1}, c2={c2}\ndebug_values={debug_values}"
                # ratio
                qb = 0.98
                qt = None
                fmt = "0.0000"
                # increase
                qpb = -0.02
                qpt = None
                fmtp = "0.000%"
            else:
                qt = 0.01
                qb = None
                fmt = "0.000000"

            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    if cell.value is None or isinstance(cell.value, str):
                        continue
                    if c1 is not None and cell.col_idx >= c1 and cell.col_idx <= c2:
                        if qt and cell.value > qpt:
                            cell.font = red
                        if qb and cell.value < qpb:
                            cell.font = red
                        cell.number_format = fmtp
                    else:
                        if qt and cell.value > qt:
                            cell.font = red
                        if qb and cell.value < qb:
                            cell.font = red
                        cell.number_format = fmt
            continue

        if k in ("bucket", "status", "op_onnx", "op_torch"):
            has_convert = [("convert" in str(c)) for c in v.columns]
            has_20 = [("-20%" in str(c)) for c in v.columns]
            assert k != "status" or any(
                has_convert
            ), f"has_convert={has_convert} but df.columns={[str(c) for c in v.columns]}"
            assert not k.startswith("bucket") or any(
                has_20
            ), f"has_20={has_20} but df.columns={[str(c) for c in v.columns]}"
            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    if cell.value is None or isinstance(cell.value, str):
                        continue
                    if cell.value == 0:
                        cell.value = None
                    cell.number_format = numbers.FORMAT_NUMBER
                    cell.alignment = center
                    if k == "status":
                        idx = cell.col_idx - n_cols - 1
                        if has_convert[idx]:
                            cell.fill = yellow
                    elif k in ("bucket",):
                        idx = cell.col_idx - n_cols - 1
                        if has_20[idx]:
                            cell.fill = redf
            continue

        if k == "time":
            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    if cell.value is None or isinstance(cell.value, str):
                        continue
                    cell.number_format = "0.000000"
            continue

        if k in {
            "SUMMARY",
            "SUMMARY2",
            "SUMMARY_base",
            "SUMMARY2_base",
            "SUMMARY2_diff",
            "SUMMARY_diff",
            "SIMPLE",
            "SIMPLE_base",
            "SIMPLE_diff",
        }:
            fmt = {
                "x": "0.000",
                "%": "0.000%",
                "bytes": "# ##0",
                "Mb": "0.000",
                "N": "0",
                "f": "0.000",
                "s": "0.0000",
                "date": "aaaa-mm-dd",
            }
            for row in sheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    if cell.value in fmt:
                        start, end = (
                            (0, cell.col_idx) if "SUMMARY" in k else (cell.col_idx, last_col)
                        )
                        for idx in range(start, end):
                            fcell = row[idx]
                            if isinstance(fcell.value, (int, float)):
                                f = fmt[cell.value]
                                if cell.value == "x":
                                    if (
                                        isinstance(fcell.value, (float, int))
                                        and fcell.value < 0.98
                                    ):
                                        fcell.font = red
                                elif cell.value in ("f", "s", "Mb"):
                                    if fcell.value >= 1000:
                                        f = "0 000"
                                    elif fcell.value >= 10:
                                        f = "0.00"
                                    elif fcell.value >= 1:
                                        f = "0.000"
                                elif cell.value == "date" and "SIMPLE" not in k:
                                    ts = time.gmtime(fcell.value)
                                    sval = time.strftime("%Y-%m-%d", ts)
                                    fcell.value = sval
                                fcell.number_format = f

            cols = {}
            for row in sheet.iter_rows(
                min_row=1,
                max_row=2,
                min_col=first_col,
                max_col=last_col,
            ):
                for cell in row:
                    if isinstance(cell.value, str):
                        cols[cell.value] = cell.col_idx

            maxc = None
            done = set()
            for k, ci in cols.items():
                if k is None or isinstance(k, int):
                    continue
                c = _column_name(ci - 1)
                if k in ("order", "#order"):
                    sheet.column_dimensions[c].width = 5
                    for cell in sheet[c]:
                        cell.alignment = right
                    done.add(c)
                    continue
                if k == "METRIC":
                    sheet.column_dimensions[c].width = 50
                    for cell in sheet[c]:
                        cell.alignment = alignment
                    done.add(c)
                    continue
                if k in {"exporter", "opt_patterns", "rtopt", "suite"} or k.startswith(
                    "version"
                ):
                    sheet.column_dimensions[c].width = 15
                    for cell in sheet[c]:
                        cell.alignment = alignment
                    done.add(c)
                    continue
                if k == "unit":
                    sheet.column_dimensions[c].width = 7
                    for cell in sheet[c]:
                        cell.alignment = right
                    maxc = c
                    done.add(c)
                    continue
                if k in ("help", "~help"):
                    sheet.column_dimensions[c].width = 20
                    for cell in sheet[c]:
                        cell.alignment = alignment
                    done.add(c)
                    continue

            for c in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
                if c == maxc:
                    break
                if c in done:
                    continue
                sheet.column_dimensions[c].width = 20
                for cell in sheet[c]:
                    cell.alignment = right
            continue

    if verbose:
        print("[_apply_excel_style] done")


def merge_benchmark_reports(
    data: Union[pandas.DataFrame, List[str], str],
    model=("suite", "model_name"),
    keys=(
        "architecture",
        "exporter",
        "opt_patterns",
        "rtopt",
        "device",
        "device_name",
        "dtype",
        "dynamic",
        "flag_fake_tensor",
        "flag_no_grad",
        "flag_training",
        "machine",
        "processor",
        "processor_name",
        "version_python",
        "version_onnx",
        "version_onnxruntime",
        "version_onnxscript",
        "version_tag",
        "version_torch",
        "version_transformers",
        "version_monai",
        "version_timm",
        "version_torch_onnx",
    ),
    column_keys=(
        "stat",
        "exporter",
        "opt_patterns",
        "rtopt",
    ),
    report_on=(
        "speedup",
        "speedup_increase",
        "speedup_med",
        "discrepancies_*",
        "TIME_ITER",
        "time_*",
        "ERR_*",
        "onnx_*",
        "op_*",
        "memory_*",
        "mem_*",
        "config_*",
    ),
    formulas=(
        "memory_peak",
        "buckets",
        "status",
        "memory_delta",
        "control_flow",
        "pass_rate",
        "accuracy_rate",
        "date",
        "correction",
    ),
    excel_output: Optional[str] = None,
    exc: bool = True,
    filter_in: Optional[str] = None,
    filter_out: Optional[str] = None,
    verbose: int = 0,
    output_clean_raw_data: Optional[str] = None,
    baseline: Optional[pandas.DataFrame] = None,
    export_simple: Optional[str] = None,
) -> Dict[str, pandas.DataFrame]:
    """
    Merges multiple files produced by bash_benchmark...

    ::

        _index,DATE,ERR_export,ITER,TIME_ITER,capability,cpu,date_start,device,device_name,...
        101Dummy-custom,2024-07-08,,0,7.119158490095288,7.0,40,2024-07-08,cuda,...
        101Dummy-script,2024-07-08,,1,6.705480073112994,7.0,40,2024-07-08,cuda,...
        101Dummy16-custom,2024-07-08,,2,6.970448340754956,7.0,40,2024-07-08,cuda,...

    :param data: dataframe with the raw data
    :param model: columns defining a unique model
    :param keys: colimns definined a unique experiment
    :param report_on: report on those metrics, ``<prefix>*`` means all
        columns starting with this prefix
    :param formulas: add computed metrics
    :param excel_output: output the computed dataframe into a excel document
    :param exc: raise exception by default
    :param filter_in: filter in some data to make the report smaller (see below)
    :param filter_out: filter out some data to make the report smaller (see below)
    :param verbose: verbosity
    :param output_clean_raw_data: output the concatenated raw data so that it can
        be used later to make a comparison
    :param baseline: to compute difference
    :param export_simple: if not None, export simple in this file.
    :return: dictionary of dataframes

    Every key with a unique value is removed.
    Every column with a unique value is displayed on main.
    List of knowns columns::

        DATE
        ERR_export
        ERR_warmup
        ITER
        TIME_ITER
        capability
        cpu
        date_start
        device
        device_name
        discrepancies_abs
        discrepancies_rel
        dtype
        dump_folder
        dynamic
        executable
        exporter
        ...

    Argument `filter_in` or `filter_out` follows the syntax
    ``<column1>:<fmt1>/<column2>:<fmt2>``.

    The format is the following:

    * a value or a set of values separated by ``;``
    """
    if baseline:
        base_dfs = merge_benchmark_reports(
            baseline,
            model=model,
            keys=keys,
            column_keys=column_keys,
            report_on=report_on,
            formulas=formulas,
            exc=exc,
            filter_in=filter_in,
            filter_out=filter_out,
            verbose=max(verbose - 1, 0),
        )
    else:
        base_dfs = None

    if isinstance(data, str):
        data = [data]

    if isinstance(data, list):
        if verbose:
            print(f"[merge_benchmark_reports] start with {len(data)} dataframes")
        dfs = []
        for filename in data:
            if isinstance(filename, str):
                try:
                    df = pandas.read_csv(filename)
                except (FileNotFoundError, pandas.errors.ParserError) as e:
                    found = glob.glob(filename)
                    if not found:
                        raise AssertionError(f"Unable to find {filename!r}") from e
                    for f in found:
                        try:
                            df = pandas.read_csv(f)
                        except pandas.errors.ParserError as ee:
                            raise AssertionError(f"Unable to read {f!r}") from ee
                        dfs.append(df)
                    continue
            elif isinstance(filename, pandas.DataFrame):
                df = filename
            else:
                raise TypeError(f"Unexpected type {type(filename)} for one element of data")
            if df.columns[0] == "#order":
                # probably a status report
                continue
            updates = {}
            for c in df.columns:
                values = set(df[c])
                if values & {True, False, np.nan} == values:
                    updates[c] = (
                        df[c]
                        .apply(lambda x: x if np.isnan(x) else (1 if x else 0))
                        .astype(float)
                    )
                if (
                    df[c].dtype not in {float, np.float64, np.float32}
                    and "_qu" not in c
                    and c.startswith(("time_", "discrepancies_", "memory"))
                ):
                    try:
                        val = df[c].astype(float)
                    except (ValueError, TypeError) as e:
                        raise AssertionError(
                            f"Unable to convert to float column {c!r} from file "
                            f"{filename!r}, values\n---\n"
                            f"{pprint.pformat(list(enumerate(zip(df['_index'], df[c]))))}"
                        ) from e
                    updates[c] = val

            if updates:
                for k, v in updates.items():
                    df[k] = v
            dfs.append(df)
        df = pandas.concat(dfs, axis=0)
    elif isinstance(data, pandas.DataFrame):
        if verbose:
            print("[merge_benchmark_reports] start with 1 dataframe")
        df = data
    else:
        raise TypeError(f"Unexpected type {type(data)} for data")

    if verbose:
        print(f"[merge_benchmark_reports] shape={df.shape}")

    if "STEP" in df.columns:
        # Experiment is run in two step. We remove the export rows
        # if it was successful as the metrics are reported in the row with the speedup.
        if verbose:
            print("[merge_benchmark_reports] remove exporter rows (column STEP)")

        if any(df["STEP"].isna()):
            vals = set(df["STEP"][~df["STEP"].isna()])
            if vals == {"last"}:
                if "ERR_export" in df.columns:
                    df = df[(df["STEP"] == "last") | ~df["ERR_export"].isna()]
                else:
                    df = df[(df["STEP"] == "last")]
            elif "ERR_export" in df.columns:
                df = df[
                    (df["STEP"].isna()) | (df["STEP"] != "export") | ~df["ERR_export"].isna()
                ]
            else:
                df = df[(df["STEP"].isna()) | (df["STEP"] != "export")]
        elif "ERR_export" in df.columns:
            df = df[(df["STEP"] != "export") | ~df["ERR_export"].isna()]
        else:
            df = df[df["STEP"] != "export"]

        if verbose:
            print(f"[merge_benchmark_reports] done, new shape={df.shape}")

    if isinstance(model, str):
        model = [model]
    elif isinstance(model, tuple):
        model = list(model)
    assert isinstance(model, list), f"Unexpected type {type(model)} for model={model}"

    # Let's rename version into version_python
    if "version" in df.columns:
        df = df.copy()
        df["version_python"] = df["version"]
        df = df.drop("version", axis=1)

    if verbose:
        print(f"[merge_benchmark_reports] model={model!r}")

    # checks all columns defining a model are available
    for m in model:
        if m not in df.columns:
            df = df.copy()
            df[m] = ""

    if filter_in or filter_out:
        if verbose:
            print("[merge_benchmark_reports] filtering data")

        df = _filter_data(df, filter_in=filter_in, filter_out=filter_out)

        if verbose:
            print(f"[merge_benchmark_reports] done, new shape={df.shape}")
        if df.shape[0] == 0:
            return {}

    if verbose:
        print("[merge_benchmark_reports] remove empty lines")
    # let's remove the empty line
    df = df[~df[model].isna().max(axis=1)].copy()

    if verbose:
        print(f"[merge_benchmark_reports] new shape={df.shape}")

    # replace nan values
    # groupby do not like nan values
    set_columns = set(df.columns)
    for c in ["opt_patterns", "rtopt", "ERR_export", "ERR_warmup"]:
        if c in set_columns:
            df[c] = df[c].fillna("-")

    # unique values
    unique = {}
    for c in df.columns:
        u = df[c].dropna().unique()
        if len(u) == 1:
            unique[c] = u.tolist()[0]
    if "exporter" in unique:
        del unique["exporter"]

    # replace nan values in key columns
    # groupby do not like nan values
    for c in keys:
        if c in set_columns:
            if verbose and c in df.columns:
                print(
                    f"[merge_benchmark_reports] KEY {len(set(df[c].dropna()))} "
                    f"unique values for {c!r} : {set(df[c].dropna())}"
                )
            if c.startswith("flag"):
                df[c] = df[c].astype(bool).fillna(False)
            elif c in {"dynamic"}:
                df[c] = df[c].fillna(0)
            elif c.startswith("version"):
                df[c] = df[c].fillna("")
            else:
                df[c] = df[c].fillna("-")

    #######################
    # preprocessing is done
    #######################
    res = {"0raw": df}

    # uniques keys
    report_on_star = [r for r in report_on if "*" in r]
    report_on = [r for r in report_on if "*" not in r]
    keys = [k for k in keys if k in set_columns]
    report_on = [k for k in report_on if k in set_columns]
    for col in sorted(set_columns):
        for star in report_on_star:
            assert star[-1] == "*", f"Unsupported position for * in {star!r}"
            s = star[:-1]
            if col.startswith(s):
                report_on.append(col)
                break

    if verbose:
        print(f"[merge_benchmark_reports] report_on {len(report_on)} metrics")

    main = [dict(column="dates", value=", ".join(sorted(df["DATE"].unique().tolist())))]
    for k, v in unique.items():
        main.append(dict(column=k, value=v))
    res["0main"] = pandas.DataFrame(main)
    new_keys = [k for k in keys if k not in unique]

    # new_keys should not be empty.
    if not new_keys:
        for m in column_keys:
            if m == "stat":
                continue
            if m in df.columns:
                new_keys.append(m)
                break
    assert new_keys, f"new_keys is empty, column_keys={column_keys}"

    # remove duplicated rows
    if verbose:
        print("[merge_benchmark_reports] remove duplicated rows")
    full_index = [s for s in df.columns if s in {*column_keys, *keys, *model}]
    dupli = ~df.duplicated(full_index, keep="last")
    df = df[dupli].copy()
    if verbose:
        print(f"[merge_benchmark_reports] done, shape={df.shape}")
        if output_clean_raw_data:
            print(
                f"[merge_benchmark_reports] save clean raw data in {output_clean_raw_data!r}"
            )
            df.to_csv(output_clean_raw_data)
        elif verbose > 1 and excel_output:
            nn = f"{excel_output}.raw.csv"
            print(f"[merge_benchmark_reports] save raw data in {nn!r}")
            df.to_csv(nn)

    # formulas
    for expr in formulas:
        if verbose:
            print(f"[merge_benchmark_reports] process formula={expr}")

        if expr == "memory_delta":
            if (
                "memory_begin" in set_columns
                and "memory_peak" in set_columns
                and "memory_end" in set_columns
            ):
                df["memory_peak_cpu_pp"] = (
                    np.maximum(df["memory_peak"], df["memory_end"]) - df["memory_begin"]
                )
                report_on.append("memory_peak_cpu_pp")
            delta_gpu = None
            m_gpu = None
            for i in range(1024):
                c1 = f"memory_gpu{i}_begin"
                c2 = f"memory_gpu{i}_peak"
                if c1 in set_columns and c2 in set_columns:
                    d = df[c2] - df[c1]
                    if delta_gpu is None:
                        delta_gpu = d
                        m_gpu = d
                    else:
                        delta_gpu += d
                        m_gpu += d
                else:
                    break
            if delta_gpu is not None:
                df["memory_peak_gpu_pp"] = m_gpu
                df["memory_delta_peak_gpu_pp"] = delta_gpu * 2**20
                report_on.append("memory_peak_gpu_pp")
                report_on.append("memory_delta_peak_gpu_pp")
            if "memory_peak_cpu_pp" in df.columns:
                df["memory_delta_peak_cpu_pp"] = df["memory_peak_cpu_pp"] * 2**20
                report_on.append("memory_delta_peak_cpu_pp")

        if expr == "memory_peak":
            if (
                "mema_gpu_5_after_export" in set_columns
                and "mema_gpu_4_reset" in set_columns
                and "mema_gpu_1_after_loading" in set_columns
                and "mema_gpu_2_after_warmup" in set_columns
                and "mema_gpu_6_before_session" in set_columns
                and "mema_gpu_8_after_export_warmup" in set_columns
            ):
                col_name = "memory_delta_peak_gpu_export"
                df[col_name] = df["mema_gpu_5_after_export"] - df["mema_gpu_4_reset"]
                report_on.append(col_name)

                col_name = "memory_peak_gpu_export"
                df[col_name] = df["mema_gpu_5_after_export"]
                report_on.append(col_name)

                col_name = "memory_delta_peak_gpu_eager_warmup"
                df[col_name] = (
                    df["mema_gpu_2_after_warmup"] - df["mema_gpu_0_before_loading"]
                )
                report_on.append(col_name)

                col_name = "memory_peak_gpu_eager_warmup"
                df[col_name] = df["mema_gpu_2_after_warmup"]
                report_on.append(col_name)

                col_name = "memory_delta_peak_gpu_warmup"
                df[col_name] = (
                    df["mema_gpu_8_after_export_warmup"] - df["mema_gpu_6_before_session"]
                )
                report_on.append(col_name)

                col_name = "memory_peak_gpu_warmup"
                df[col_name] = df["mema_gpu_8_after_export_warmup"]
                report_on.append(col_name)
            continue

        if expr == "pass_rate":
            if "discrepancies_abs" in set_columns and "speedup" in set_columns:
                col = (df["discrepancies_abs"] <= 0.1) & (df["speedup"] >= 0.98)
                df["status_pass_rate"] = col.astype(int)
                df.loc[df["discrepancies_abs"].isna(), "status_pass_rate"] = np.nan
                report_on.append("status_pass_rate")
            continue

        if expr == "correction":
            if "time_latency_eager" in df.columns and "time_latency" in df.columns:
                weights = df["time_latency"].apply(lambda x: np.nan if np.isnan(x) else 1.0)
                df["time_latency_eager_if_ort"] = df["time_latency_eager"] * weights
                report_on.append("time_latency_eager_if_ort")
            continue

        if expr == "accuracy_rate":
            if "discrepancies_abs" in set_columns:
                col = df["discrepancies_abs"] <= 0.1
                df["status_accuracy_rate"] = col.astype(int)
                df.loc[df["discrepancies_abs"].isna(), "status_accuracy_rate"] = np.nan
                report_on.append("status_accuracy_rate")
            continue

        if expr == "date":
            if "date_start" in set_columns:
                df["status_date"] = (
                    pandas.to_datetime(df["date_start"]).astype("int64").astype(float) / 1e9
                )
                set_columns = set(df.columns)
                report_on.append("status_date")
            continue

        if expr == "status":
            if "time_export_success" in set_columns:
                df["status_convert"] = (~df["time_export_success"].isna()).astype(int)
                report_on.append("status_convert")
            if "discrepancies_abs" in set_columns:
                df["status_convert_ort"] = (~df["discrepancies_abs"].isna()).astype(int)
                mets = []
                for th, mt in itertools.product(
                    ["1e-1", "1e-2", "1e-3", "1e-4"], ["abs", "abs_0", "abs_1+"]
                ):
                    dis = f"discrepancies_{mt}"
                    if dis not in df.columns:
                        continue
                    met = f"status_err{mt[3:]}<{th}"
                    mets.append(met)
                    df[met] = (~df[dis].isna() & (df[dis] < float(th))).astype(int)
                df["status_lat<=eager+2%"] = (
                    ~df["discrepancies_abs"].isna()
                    & (df["time_latency"] <= df["time_latency_eager"] * 1.02)
                ).astype(int)
                set_columns = set(df.columns)
                report_on.extend(
                    [
                        "status_convert_ort",
                        *mets,
                        "status_lat<=eager+2%",
                    ]
                )
            continue

        if expr == "control_flow":
            if (
                "exporter" in set_columns
                and "time_export_success" in set_columns
                and ({"export", "compile"} & set(df.exporter))
                and len(set(df.exporter)) > 1
            ):
                expo = "export" if "export" in set(df.exporter) else "compile"
                keep = [*model, *new_keys, "time_export_success"]
                gr = df[df.exporter == expo][keep].copy()
                gr["status_control_flow"] = gr["time_export_success"].isna().astype(int)
                gr = gr.drop("time_export_success", axis=1)
                if "opt_patterns" in gr.columns and len(set(gr.opt_patterns)) == 1:
                    on = [
                        k
                        for k in keep[:-1]
                        if k not in ("exporter", "opt_patterns", "rtopt")
                    ]
                else:
                    on = [k for k in keep[:-1] if k != "exporter"]
                joined = pandas.merge(df, gr, left_on=on, right_on=on, how="left")
                assert (
                    df.shape[0] == joined.shape[0]
                ), f"Shape mismatch after join {df.shape} -> {joined.shape}"
                df = joined.copy()
                for c in column_keys:
                    cc = f"{c}_x"
                    if cc in df.columns:
                        df[c] = df[cc]
                drop = [
                    c
                    for c in [
                        *[f"{c}_x" for c in column_keys],
                        *[f"{c}_y" for c in column_keys],
                    ]
                    if c in df.columns
                ]
                df = df.drop(drop, axis=1)
                set_columns = set(df.columns)
                report_on.append("status_control_flow")
            continue

        if expr == "buckets":
            if "rtopt" not in df.columns:
                # We assume onnxruntime optimization were enabled (default settings).
                df["rtopt"] = 1
            if (
                "exporter" in set_columns
                and "opt_patterns" in set_columns
                and "speedup" in set_columns
                and "torch_script" in set(df.exporter)
                and len(set(df.exporter)) > 1
            ):
                # Do the same with the exporter as a baseline.
                keep = [*model, *new_keys, "speedup"]
                gr = df[
                    (df.exporter == "torch_script")
                    & (df.opt_patterns.isin(["", "-"]))
                    & (df.rtopt == 1)
                ][keep].copy()
                gr = gr[~gr["speedup"].isna()]

                if gr.shape[0] == 0:
                    # No figures for torch_script
                    if verbose:
                        print(
                            f"[merge_benchmark_reports] gr.shape={gr.shape}, "
                            f"unable to compute speedup_script, "
                            f"exporters={set(df.exporter)}, "
                            f"opt_patterns={set(df.opt_patterns)}, rtopt={set(df.rt_opt)}"
                        )
                else:
                    if verbose:
                        print(
                            f"[merge_benchmark_reports] compute setup_script: "
                            f"gr.shape={gr.shape}"
                        )
                    gr["speedup_script"] = gr["speedup"]
                    gr = gr.drop("speedup", axis=1)

                    on = [
                        k
                        for k in keep[:-1]
                        if k not in ("exporter", "opt_patterns", "rtopt")
                    ]
                    joined = pandas.merge(df, gr, left_on=on, right_on=on, how="left")

                    assert df.shape[0] == joined.shape[0], (
                        f"Shape mismatch after join {df.shape} -> {joined.shape}, "
                        f"gr.shape={gr.shape}, on={on}"
                    )
                    df = joined.copy()
                    df["speedup_increase_script"] = (
                        df["speedup"] / df["speedup_script"] - 1
                    ).fillna(-np.inf)
                    report_on.extend(["speedup_script", "speedup_increase_script"])
                    for c in column_keys:
                        cc = f"{c}_x"
                        if cc in df.columns:
                            df[c] = df[cc]
                    drop = [
                        c
                        for c in [
                            *[f"{c}_x" for c in column_keys],
                            *[f"{c}_y" for c in column_keys],
                        ]
                        if c in df.columns
                    ]
                    df = df.drop(drop, axis=1)
                    set_columns = set(df.columns)
                    df["status_lat<=script+2%"] = (
                        df["speedup_increase_script"] >= (1 / 1.02 - 1)
                    ).astype(int)
                    report_on.append("status_lat<=script+2%")

            for c in ["speedup_increase", "speedup_increase_script"]:
                if c not in set_columns:
                    continue
                scale = BUCKET_SCALES
                ind = df["speedup_increase"].isna()
                for i in range(1, len(scale)):
                    val = (df[c] >= scale[i - 1] / 100) & (df[c] < scale[i] / 100)
                    v1 = f"{scale[i-1]}%" if not np.isinf(scale[i - 1]) else ""
                    v2 = f"{scale[i]}%" if not np.isinf(scale[i]) else ""
                    suf = f"[{v1},{v2}[" if v1 and v2 else (f"<{v2}" if v2 else f">={v1}")
                    if c == "speedup_increase_script":
                        d = f"bucket_script {suf}"
                    else:
                        d = f"bucket_{suf}"
                    df[d] = val.astype(int)
                    df.loc[ind, d] = np.nan
                    report_on.append(d)
            continue

    if verbose:
        print(f"[merge_benchmark_reports] done, shape={df.shape}")

    # values
    for c in report_on:
        keep = [*model, *new_keys, c]
        dfc = df[keep]
        dfc = dfc[~dfc[model].isna().min(axis=1)]
        if new_keys:
            pivot = dfc.pivot(index=model, columns=new_keys, values=c)
        else:
            pivot = dfc.set_index(model)
        res[c] = pivot

    if verbose:
        print(f"[merge_benchmark_reports] {len(res)} metrics")

    # let's remove empty variables
    if verbose:
        print("[merge_benchmark_reports] remove empty variables")

    for v in res.values():
        drop = []
        for c in v.columns:
            if all(v[c].isna()) or set(v[c]) == {"-"}:
                drop.append(c)
        if drop:
            v.drop(drop, axis=1, inplace=True)
    res = {k: v for k, v in res.items() if v.shape[1] > 0}
    if "TIME_ITER" in res:
        res["time_ITER"] = res["TIME_ITER"]
        del res["TIME_ITER"]

    if verbose:
        print(f"[merge_benchmark_reports] final number of metrics={len(res)}")

    # final fusion

    def _merge(res, merge, prefix, reverse=True, transpose=False):
        m = None
        for name in merge:
            df = res[name].T
            index_cols = df.index.names
            if index_cols == [None]:
                index_cols = ["index"]
            df["stat"] = name[len(prefix) :]
            df = df.reset_index(drop=False).set_index([*list(index_cols), "stat"])
            # Let's remove duplicated experiment, the last one is kept.
            df = df.T
            df = df[~df.index.duplicated(keep="last")].copy()
            if m is None:
                m = df
                continue
            m0 = m
            try:
                m = pandas.merge(
                    m,
                    df,
                    how="outer",
                    left_index=True,
                    right_index=True,
                    validate="1:1",
                )
            except ValueError as e:
                raise AssertionError(
                    f"Unable to join for name={name}, df.index={df.index.names}, "
                    f"m.index={m.index.names} e={e}\n----\n+ m0.index\n{m0.index[:3]}"
                    f"\n---\n+ df.index\n{df.index[:3]}"
                ) from e
            assert m.shape[0] <= df.shape[0] + m0.shape[0], (
                f"Something is going wrong for prefix {prefix!r} "
                f"(probably a same experiment reported twice), "
                f"df.shape={df.shape}, new_shape is {m.shape} "
                f"(before shape={m0.shape})"
                f"\n-- m0=\n{m0}\n-- df=\n{df}"
            )

        # We need to change the columns index order.
        if reverse:
            df = m.T
            index = set(df.index.names)
            if index == {"stat", "exporter"}:
                m = df.reset_index(drop=False).set_index(["stat", "exporter"]).T
            elif index == {"stat", "index"}:
                m = (
                    df.reset_index(drop=False)
                    .sort_index(axis=1)
                    .drop("index", axis=1)
                    .set_index(["stat"])
                    .T
                )
        else:
            m = m.T.sort_index()
            m = m.T
            if list(m.columns.names) == ["index", "stat"]:
                m.columns = m.columns.droplevel(level=0)

        if transpose:
            m0 = m
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=FutureWarning)
                m = m.T.stack(level=list(range(len(m.index.names)))).reset_index(drop=False)
            cols = m.columns
            assert len(cols) >= 4, (
                f"Unexpected number of columns in {cols}, "
                f"prefix={prefix!r}, m.columns={m.columns}, "
                f"m0.index.names={m0.index.names}, "
                f"m0.columns.names={m0.columns.names}\n---\n{m0}"
            )
            exporter_column = [c for c in cols if c in ("exporter", "opt_patterns", "rtopt")]
            assert "stat" in cols, (
                f"Unexpected columns {cols}, expecting 'stat', "
                f"{exporter_column!r}, {model!r}, reverse={reverse}, "
                f"transpose={transpose}, m0.index.names={m0.index.names}, "
                f"m0.columns.names={m0.columns.names}\n---"
                f"\n{m0.head()}\n---\n{m.head()}"
            )
            last = [c for c in cols if c not in {"stat", *exporter_column, *model}]
            assert last, (
                f"last cannot be empty, exporter_column={exporter_column}, "
                f"model={model}, cols={cols}\n----\n{m0.head()}"
            )
            added_columns = [c for c in last if c in new_keys]
            last = [c for c in last if c not in new_keys]
            if len(last) == 2 and last[0] == "index":
                last = last[1:]
            assert len(last) == 1, (
                f"Unexpected columns in {cols}, added={added_columns}, "
                f"last={last}, new_keys={new_keys}, exporter_column={exporter_column}, "
                f"prefix={prefix!r}"
                f"\nm0.index.names={m0.index.names}"
                f"\nm0.columns.names={m0.columns.names}"
                f"\nm0.columns[0]={m0.columns[0]}"
                f"\nm0.columns={m0.columns}\n----"
                f"\nm.index.names={m.index.names}"
                f"\nm.columns.names={m.columns.names}"
                f"\nm.columns[0]={m.columns[0]}"
                f"\nm.columns={m.columns}\n----"
            )
            m = m.pivot(
                index="stat",
                columns=[*model, *exporter_column, *added_columns],
                values=last[0],
            )
            m = m.T.sort_index().T
        return m

    if "speedup" in res:
        res["speedup_1speedup"] = res["speedup"]
        del res["speedup"]

    # verification

    set_model = set(model)
    for c, v in res.items():
        assert (
            c in {"0raw", "0main"} or set_model & set(v.index.names) == set_model
        ), f"There should not be any multiindex but c={c!r}, names={v.index.names}"

    # MODELS
    res["MODELS"] = _select_model_metrics(
        res,
        select=SELECTED_FEATURES,
        stack_levels=tuple(c for c in column_keys if c != "stat"),
    )

    # merging

    for prefix in [
        "status_",
        "discrepancies_",
        "memory_",
        "onnx_",
        "time_",
        "ERR_",
        "op_onnx_",
        "op_torch_",
        "mempeak_",
        "speedup_",
        "bucket_",
        "config_",
    ]:
        merge = [k for k in res if k.startswith(prefix)]
        merge.sort()

        if len(merge) == 0:
            continue

        if verbose:
            print(f"[merge_benchmark_reports] start fusion of {prefix!r}")

        sheet = _merge(
            res,
            merge,
            prefix,
            reverse=prefix != "status_",
            transpose=prefix.startswith("op_"),
        )
        assert (
            None not in sheet.index.names
        ), f"None in sheet.index.names={sheet.index.names}, prefix={prefix!r}"
        assert (
            None not in sheet.columns.names
        ), f"None in sheet.columns.names={sheet.columns.names}, prefix={prefix!r}"
        if verbose:
            print(f"[merge_benchmark_reports] done, shape of {prefix!r} is {sheet.shape}")

        res[prefix[:-1]] = sheet
        res = {k: v for k, v in res.items() if k not in set(merge)}

    # try to use numerical value everywhere
    if verbose:
        print("[merge_benchmark_reports] enforce numerical values")
    for k, v in res.items():
        if k in {"0main"}:
            continue
        for c in v.columns:
            if "output_names" in c or "input_names" in c:
                continue
            if "date" in c:
                continue
            cc = v[c]
            if cc.dtype == np.object_:
                try:
                    dd = cc.astype(float)
                    v[c] = dd
                except (ValueError, TypeError):
                    types = [
                        type(_) for _ in cc if not isinstance(_, float) or not np.isnan(_)
                    ]
                    if set(types) == {str}:
                        continue
                    co = Counter(types)
                    assert str in co and co[str] >= len(types) // 4 + 1, (
                        f"Unexpected values for k={k!r}, columns={c!r}, "
                        f"types={set(types)}, values={v[c]}, co={co}"
                    )

    if verbose:
        print("[merge_benchmark_reports] done")
        print("[merge_benchmark_reports] create aggregated figures")

    # add pages for the summary
    res["AGG"], res["AGG2"] = _create_aggregation_figures(
        res,
        skip={
            "ERR",
            "0raw",
            "0main",
            "op_onnx",
            "op_torch",
            "MODELS",
        },
        model=model,
        exc=exc,
    )
    assert (
        None not in res["AGG"].index.names
    ), f"None in res['AGG'].index.names={res['AGG'].index.names}, prefix='AGG'"
    assert (
        None not in res["AGG2"].index.names
    ), f"None in res['AGG2'].index.names={res['AGG2'].index.names}, prefix='AGG2'"
    if verbose:
        print(
            f"[merge_benchmark_reports] done with shapes "
            f"{res['AGG'].shape} and {res['AGG2'].shape}"
        )

    names = res["AGG"].index.names
    new_names = []
    for c in ["cat", "stat"]:
        if c in names:
            new_names.append(c)
    last_names = []
    for c in column_keys:
        if c == "stat" or c not in names:
            continue
        last_names.append(c)
    if "agg" in names:
        last_names.append("agg")
    for c in names:
        if c not in new_names and c not in last_names:
            new_names.append(c)

    if verbose:
        print("[merge_benchmark_reports] reorder")

    res["AGG"] = _reorder_index_level(res["AGG"], new_names + last_names, prefix="AGG")
    res["AGG2"] = _reorder_index_level(res["AGG2"], new_names + last_names, prefix="AGG2")

    if verbose:
        print("[merge_benchmark_reports] done")

    final_res = {
        k: (
            v
            if k in {"0raw", "0main", "AGG", "AGG2", "op_onnx", "op_torch"}
            else _reorder_columns_level(v, column_keys, prefix=k)
        )
        for k, v in res.items()
    }

    if verbose:
        print(f"[merge_benchmark_reports] done with {len(final_res)} sheets")
        print("[merge_benchmark_reports] creates SUMMARY, SUMMARY2, SIMPLE")

    final_res["SUMMARY"], _suites = _select_metrics(
        res["AGG"], select=SELECTED_FEATURES, prefix="SUMMARY"
    )
    final_res["SIMPLE"], _suites = _select_metrics(
        res["AGG"],
        select=[f for f in SELECTED_FEATURES if f.get("simple", False)],
        prefix="SIMPLE",
    )
    final_res["SUMMARY2"], _suites = _select_metrics(
        res["AGG2"], select=SELECTED_FEATURES, prefix="SUMMARY2"
    )

    # adding dates
    df0 = final_res["0raw"]
    date_col = [
        c
        for c in ["DATE", "suite", "exporter", "opt_patterns", "dtype", "rtopt"]
        if c in df0.columns
    ]
    if verbose:
        print(f"[merge_benchmark_reports] add dates with columns={date_col}")
    date_col2 = [c for c in date_col if c != "DATE"]
    assert len(date_col2) != len(date_col), f"No date found in {sorted(df0.columns)}"
    final_res["dates"] = df0[date_col].groupby(date_col2).max().reset_index(drop=False)
    date_col2 = [c for c in date_col2 if c in final_res["SIMPLE"]]
    assert "suite" in date_col2, f"Unable to find 'suite' in {date_col2}"
    simple = final_res["SIMPLE"].merge(
        final_res["dates"], left_on=date_col2, right_on=date_col2, how="left"
    )
    assert (
        simple.shape[0] == final_res["SIMPLE"].shape[0]
    ), f"Some rows were added or deleted {final_res['SIMPLE'].shape} -> {simple.shape}"
    final_res["SIMPLE"] = simple

    if verbose:
        print(
            f"[merge_benchmark_reports] done with shapes "
            f"{final_res['SUMMARY'].shape}, {final_res['SUMMARY2'].shape}, "
            f"{final_res['SIMPLE'].shape}"
        )

    if base_dfs:

        def _float_(x):
            if isinstance(x, (int, float)):
                return x
            try:
                return float(x)
            except (ValueError, TypeError):
                return np.nan

        for name in {"0main", "0raw", "SUMMARY", "SUMMARY2", "MODELS", "SIMPLE"}:
            if name in base_dfs:
                final_res[f"{name}_base"] = base_dfs[name]

        for name in {"SUMMARY", "SUMMARY2", "MODELS", "SIMPLE"}:
            if name not in base_dfs or name not in final_res:
                continue
            drop = []
            df_str = final_res[name].select_dtypes("object")
            if df_str.shape[1] > 0:
                # The date may set the type of object just for one value
                # Let cast every columns to see if it can improved.
                df_base = base_dfs[name].copy()
                for c in df_base.columns:
                    cc = df_base[c].apply(_float_).astype(float)
                    if cc.isna().sum() <= 2:
                        df_base[c] = cc
                df_this = final_res[name].copy()
                for c in df_this.columns:
                    cc = df_this[c].apply(_float_).astype(float, errors="ignore")
                    if cc.isna().sum() <= 2:
                        df_this[c] = cc
                df_str = df_this.select_dtypes("object")
            else:
                df_this = final_res[name]
                df_base = base_dfs[name]

            if df_str.shape[1] > 0:
                stacked = list(df_str.columns)
                order_name = [c for c in df_this.columns if "#order" in c]
                if len(order_name) == 1:
                    stacked.extend(order_name)
                    drop.extend(order_name)
                assert all((c in df_this and c in df_base) for c in df_str.columns), (
                    f"Columns mismatch in sheet {name!r}, "
                    f"columns={list(df_str.columns)},\n"
                    f"[{[int(c in df_this) for c in df_str.columns]}], "
                    f"[{[int(c in df_base) for c in df_str.columns]}], "
                    f"\n{list(df_this.columns)}\n!=\n{list(df_base.columns)}"
                )
                df_this = df_this.set_index(stacked)
                df_base = df_base.set_index(stacked)
            else:
                stacked = None

            df_this = df_this.select_dtypes("number")
            df_base = df_base.select_dtypes("number")
            set_columns = sorted(set(df_this.columns) & set(df_base.columns))
            df_base = df_base[set_columns].sort_index(axis=0).sort_index(axis=1)
            df_this = df_this[set_columns].sort_index(axis=0).sort_index(axis=1)
            df_num = df_this.sub(df_base)
            if stacked:
                df_num = df_num.reset_index(drop=False)
                order_name = [c for c in df_num.columns if "#order" in c]
                if len(order_name) == 1:
                    df_num = df_num.sort_values(order_name)

            final_res[f"{name}_diff"] = df_num.sort_index(axis=1)

    # cleaning empty columns
    for v in final_res.values():
        v.dropna(axis=1, how="all", inplace=True)

    if export_simple and "SIMPLE" in final_res:
        if (
            "rtopt" in final_res["SIMPLE"].columns
            and len(set(final_res["SIMPLE"]["rtopt"].dropna())) <= 1
        ):
            if verbose:
                print("[merge_benchmark_reports] drops 'rtopt' in SIMPLE")
            final_res["SIMPLE"] = final_res["SIMPLE"].drop("rtopt", axis=1)
        elif verbose and "rtopt" in final_res["SIMPLE"].columns:
            print(
                f"[merge_benchmark_reports] keeps 'rtopt' in SIMPLE: "
                f"{set(final_res['SIMPLE'].dropna())}"
            )
        if verbose:
            print(f"[merge_benchmark_reports] writes {export_simple!r}")

        final_res["SIMPLE"].to_csv(export_simple, index=False)
        _set_ = set(final_res["SIMPLE"])

        # first pivot
        piv_index = tuple(c for c in ("dtype", "suite", "#order", "METRIC") if c in _set_)
        piv_columns = tuple(c for c in ("exporter", "opt_patterns", "rtopt") if c in _set_)
        ccc = [*piv_index, *piv_columns]
        gr = final_res["SIMPLE"][[*ccc, "value"]].groupby(ccc).count()
        assert gr.values.max() <= 1, (
            f"Unexpected duplicated, piv_index={piv_index}, "
            f"piv_columns={piv_columns}, columns={final_res['SIMPLE'].columns}, "
            f"issue=\n{gr[gr['value'] > 1]}"
        )

        piv = (
            final_res["SIMPLE"]
            .pivot(
                index=piv_index,
                columns=piv_columns,
                values="value",
            )
            .sort_index(axis=0)
            .sort_index(axis=1)
        )
        export_simple_x = f"{export_simple}.xlsx"
        if verbose:
            print(f"[merge_benchmark_reports] writes {export_simple_x!r}")
        piv.to_excel(export_simple_x)

        # total
        piv_index = tuple(c for c in ("#order", "METRIC") if c in _set_)
        piv_columns = (c for c in ("exporter", "opt_patterns", "rtopt") if c in _set_)
        piv = (
            pandas.pivot_table(
                final_res["SIMPLE"],
                index=piv_index,
                columns=piv_columns,
                values="value",
                aggfunc="sum",
            )
            .sort_index(axis=0)
            .sort_index(axis=1)
        )
        export_simple_x = f"{export_simple}_total.xlsx"
        if verbose:
            print(f"[merge_benchmark_reports] writes {export_simple_x!r}")
        piv.to_excel(export_simple_x)

    if excel_output:
        if verbose:
            print(f"[merge_benchmark_reports] apply Excel style with {excel_output!r}")
        with pandas.ExcelWriter(excel_output) as writer:
            no_index = {
                "0raw",
                "0main",
                "SUMMARY",
                "SUMMARY_base",
                "SIMPLE",
                "SIMPLE_base",
            }
            first_sheet = ["0main"]
            last_sheet = [
                "ERR",
                "SUMMARY",
                "SUMMARY_base",
                "SUMMARY_diff",
                "AGG",
                "AGG2",
                "0raw",
            ]
            ordered = set(first_sheet) | set(last_sheet)
            order = [
                *first_sheet,
                *sorted([k for k in final_res if k not in ordered]),
                *last_sheet,
            ]
            order = [k for k in order if k in final_res]
            for k in order:
                v = final_res[k]
                ev = _reverse_column_names_order(v, name=k)
                frow = len(ev.columns.names) if isinstance(ev.columns.names, list) else 1
                if k.startswith("SUMMARY"):
                    fcol = len(v.columns.names)
                else:
                    fcol = len(ev.index.names) if isinstance(ev.index.names, list) else 1

                if (
                    k in {"AGG2", "SUMMARY2", "SUMMARY2_base", "SUMMARY2_diff"}
                    and "suite" in ev.columns.names
                ):
                    if verbose:
                        print(f"[merge_benchmark_reports] reorder {k!r} (1)")
                    ev = _reorder_columns_level(ev, first_level=["suite"], prefix=k)
                elif k == "MODELS":
                    if verbose:
                        print(f"[merge_benchmark_reports] reorder {k!r} (2)")
                    ev = _reorder_columns_level(ev, first_level=["#order"], prefix=k)
                if verbose:
                    print(f"[merge_benchmark_reports] add {k!r} to {excel_output!r}")
                ev.to_excel(
                    writer,
                    sheet_name=k,
                    index=k not in no_index,
                    freeze_panes=(frow, fcol) if k not in no_index else None,
                )
                if verbose:
                    print(f"[merge_benchmark_reports] added {k!r} to {excel_output!r}")
            _apply_excel_style(final_res, writer, verbose=verbose)
            if verbose:
                print(f"[merge_benchmark_reports] save in {excel_output!r}")

    return final_res


def _reorder_columns_level(
    df: pandas.DataFrame,
    first_level: List[str],
    prefix: Optional[str] = None,
) -> pandas.DataFrame:
    assert _nonone_(df.columns), f"None in {df.columns.names}, prefix={prefix!r}"
    assert set(df.columns.names) & set(first_level), (
        f"Nothing to sort, prefix={prefix!r} "
        f"df.columns={df.columns}, first_level={first_level}\n--\n{df}"
    )

    c_in = [c for c in first_level if c in set(df.columns.names)]
    c_out = [c for c in df.columns.names if c not in set(c_in)]
    levels = c_in + c_out
    for i in range(len(levels)):
        if levels[i] == df.columns.names[i]:
            continue
        j = list(df.columns.names).index(levels[i])
        df = df.swaplevel(i, j, axis=1)
    assert (
        list(df.columns.names) == levels
    ), f"Issue levels={levels}, df.columns.names={df.columns.names}"
    assert _nonone_(df.columns), f"None in {df.columns.names}"
    return df.sort_index(axis=1)


def _sort_index_level(
    df: pandas.DataFrame,
    debug: Optional[str] = None,
) -> pandas.DataFrame:
    assert _nonone_(df.index), f"None in {df.index.names}, debug={debug!r}"
    assert df.columns.names == [None] or _nonone_(
        df.columns
    ), f"None in {df.columns.names}, debug={debug!r}"
    levels = list(df.index.names)
    levels.sort()
    for i in range(len(levels)):
        if levels[i] == df.index.names[i]:
            continue
        j = list(df.index.names).index(levels[i])
        df = df.swaplevel(i, j, axis=0)
    assert (
        list(df.index.names) == levels
    ), f"Issue levels={levels}, df.index.names={df.index.names}, debug={debug!r}"
    assert _nonone_(df.index), f"None in {df.index.names}, debug={debug!r}"
    assert df.columns.names == [None] or _nonone_(
        df.columns
    ), f"None in {df.columns.names}, debug={debug!r}"
    return df.sort_index(axis=0)


def _reorder_index_level(
    df: pandas.DataFrame,
    first_level: List[str],
    prefix: Optional[str] = None,
) -> pandas.DataFrame:
    assert _nonone_(df.index), f"None in {df.index.names}, prefix={prefix!r}"
    assert _nonone_(df.columns), f"None in {df.columns.names}, prefix={prefix!r}"
    assert set(df.index.names) & set(first_level), (
        f"Nothing to sort, prefix={prefix!r} "
        f"df.columns={df.index}, first_level={first_level}"
    )

    c_in = [c for c in first_level if c in set(df.index.names)]
    c_out = [c for c in df.index.names if c not in set(c_in)]
    levels = c_in + c_out
    for i in range(len(levels)):
        if levels[i] == df.index.names[i]:
            continue
        j = list(df.index.names).index(levels[i])
        df = df.swaplevel(i, j, axis=0)
    assert (
        list(df.index.names) == levels
    ), f"Issue levels={levels}, df.columns.names={df.index.names}"
    assert _nonone_(df.index), f"None in {df.index.names}"
    assert _nonone_(df.columns), f"None in {df.columns.names}"
    return df.sort_index(axis=0)


def _add_level(
    index: pandas.MultiIndex,
    name: str,
    value: str,
) -> pandas.MultiIndex:
    if len(index.names) == 1:
        v = index.tolist()
        nv = [[value, _] for _ in v]
        nn = [name, index.names[0]]
        aa = np.array(nv).T.tolist()
        new_index = pandas.MultiIndex.from_arrays(aa, names=nn)
        return new_index

    v = index.tolist()
    nv = [[value, *_] for _ in v]
    nn = [name, *index.names]
    aa = np.array(nv).T.tolist()
    new_index = pandas.MultiIndex.from_arrays(aa, names=nn)
    return new_index


def _create_aggregation_figures(
    final_res: Dict[str, pandas.DataFrame],
    model: List[str],
    skip: Optional[Set[str]] = None,
    key: str = "suite",
    exc: bool = True,
) -> Dict[str, pandas.DataFrame]:
    assert key in model, f"Key {key!r} missing in model={model!r}"
    model_not_key = [c for c in model if c != key]

    aggs = {}
    for k, v in final_res.items():
        if k in skip:
            continue
        assert (
            v.select_dtypes(include=np.number).shape[1] > 0
        ), f"No numeric column for k={k!r}, dtypes=\n{v.dtypes}"
        assert None not in v.index.names, f"None in v.index.names={v.index.names}, k={k!r}"
        assert (
            None not in v.columns.names
        ), f"None in v.columns.names={v.columns.names}, k={k!r}"

        if key not in v.index.names:
            v = v.copy()
            v[key] = "?"
            v = v.reset_index(drop=False).set_index([key, *v.index.names])

        assert (
            key in v.index.names
        ), f"Unable to find key={key} in {v.index.names} for k={k!r}"
        assert len(v.index.names) == len(
            model
        ), f"Length mismatch for k={k!r}, v.index.names={v.index.names}, model={model}"

        # Let's drop any non numerical features.
        v = v.select_dtypes(include=[np.number])
        # gv = v.apply(lambda x: np.log(np.maximum(x, 1e-10).values))
        v = v.reset_index(drop=False).set_index(model_not_key)

        assert key in v.columns, f"Unable to find column {key!r} in {v.columns}"

        v = v.sort_index(axis=1)

        assert None not in v.index.names, f"None in v.index.names={v.index.names}, k={k!r}"
        assert (
            None not in v.columns.names
        ), f"None in v.columns.names={v.columns.names}, k={k!r}"
        try:
            gr = v.groupby(key)
        except ValueError as e:
            raise AssertionError(
                f"Unable to grouby by key={key!r}, "
                f"shape={v.shape} v.columns={v.columns}, "
                f"values={set(v[key])}\n---\n{v}"
            ) from e

        def _geo_mean(serie):
            nonan = serie.dropna()
            if len(nonan) == 0:
                return np.nan
            res = np.exp(np.log(np.maximum(nonan, 1e-10)).mean())
            return res

        def _propnan(df, is_nan):
            df[is_nan] = np.nan
            return df

        gr_no_nan = v.fillna(0).groupby(key)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=(FutureWarning, PerformanceWarning))
            total = gr_no_nan.count()
            is_nan = gr_no_nan.count() - gr.count() == total
        stats = [
            ("MEAN", gr.mean()),
            ("MEDIAN", gr.median()),
            ("SUM", _propnan(gr.sum(), is_nan)),
            ("MIN", gr.min()),
            ("MAX", gr.max()),
            ("COUNT", _propnan(gr.count(), is_nan)),
            ("COUNT%", _propnan(gr.count() / total, is_nan)),
            ("TOTAL", _propnan(total, is_nan)),
            ("NAN", _propnan(is_nan.astype(int), is_nan)),
        ]

        if k.startswith("speedup"):
            try:
                geo_mean = gr.agg(_geo_mean)
            except ValueError as e:
                if exc:
                    raise AssertionError(
                        f"Fails for geo_mean, k={k!r}, v=\n{v.head().T}"
                    ) from e
                geo_mean = None
            if geo_mean is not None:
                stats.append(("GEO-MEAN", geo_mean))

        dfs = []
        for name, df in stats:
            # avoid date to be numbers
            updates = {}
            drops = []
            for col in df.columns:
                if (isinstance(col, tuple) and "date" in col) or col == "date":
                    if name in {"SUM", "MEDIAN"}:
                        drops.append(col)
                    elif name in {"MIN", "MAX", "MEAN", "MEDIAM"}:
                        # maybe this code will fail someday but it seems that the cycle
                        # date -> int64 -> float64 -> int64 -> date
                        # keeps the date unchanged
                        vvv = df[col]
                        if vvv.dtype not in {object, np.object_, str, np.str_}:
                            updates[col] = vvv.apply(
                                lambda d: (
                                    np.nan
                                    if np.isnan(d)
                                    else time.strftime("%Y-%m-%d", time.gmtime(d))
                                )
                            )
            if drops:
                df.drop(drops, axis=1, inplace=True)
            if updates:
                for kc, vc in updates.items():
                    df[kc] = vc

            # then continue
            assert isinstance(
                df, pandas.DataFrame
            ), f"Unexpected type {type(df)} for k={k!r} and name={name!r}"
            df.index = _add_level(df.index, "agg", name)
            df.index = _add_level(df.index, "cat", k)
            assert _nonone_(df.index), f"None in {df.index.names}, k={k!r}, name={name!r}"
            assert _nonone_(
                df.columns
            ), f"None in {df.columns.names}, k={k!r}, name={name!r}"
            dfs.append(df)

        if len(dfs) == 0:
            continue
        df = pandas.concat(dfs, axis=0)

        assert df.shape[0] > 0, f"Empty set for k={k!r}"
        assert df.shape[1] > 0, f"Empty columns for k={k!r}"
        assert _nonone_(df.index), f"None in {df.index.names}, k={k!r}"
        assert _nonone_(df.columns), f"None in {df.columns.names}, k={k!r}"
        assert isinstance(df, pandas.DataFrame), f"Unexpected type {type(df)} for k={k!r}"

        if "stat" in df.columns.names:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=(FutureWarning, PerformanceWarning))
                df = df.stack("stat")
            if not isinstance(df, pandas.DataFrame):
                assert (
                    "opt_patterns" not in df.index.names
                ), f"Unexpected names for df.index.names={df.index.names} (k={k!r})"
                assert (
                    "rtopt" not in df.index.names
                ), f"Unexpected names for df.index.names={df.index.names} (k={k!r})"
                df = df.to_frame()
                if df.columns.names == [None]:
                    df.columns = pandas.MultiIndex.from_arrays(
                        [("_dummy_",)], names=["_dummy_"]
                    )
                    assert _nonone_(
                        df.columns
                    ), f"None in {df.columns.names}, k={k!r}, df={df}"
            assert isinstance(
                df, pandas.DataFrame
            ), f"Unexpected type {type(df)} for k={k!r}"
            assert _nonone_(df.index), f"None in {df.index.names}, k={k!r}"
            assert _nonone_(df.columns), f"None in {df.columns.names}, k={k!r}, df={df}"
        assert isinstance(df, pandas.DataFrame), f"Unexpected type {type(df)} for k={k!r}"
        assert _nonone_(df.index), f"None in {df.index.names}, k={k!r}"
        assert _nonone_(df.columns), f"None in {df.columns.names}, k={k!r}"
        aggs[f"agg_{k}"] = df

    # check stat is part of the column otherwise the concatenation fails

    set_names = set()
    for df in aggs.values():
        set_names |= set(df.index.names)

    for k, df in aggs.items():
        assert _nonone_(df.index), f"None in {df.index.names}, k={k!r}"
        assert _nonone_(df.columns), f"None in {df.columns.names}, k={k!r}"
        if len(df.index.names) == len(set_names):
            continue
        missing = set_names - set(df.index.names)
        for g in missing:
            df.index = _add_level(df.index, g, k.replace("agg_", ""))

    aggs = {k: _sort_index_level(df, debug=k) for k, df in aggs.items()}

    # concatenation
    dfs = pandas.concat(list(aggs.values()), axis=0)
    assert None not in dfs.index.names, f"None in dfs.index.names={dfs.index.names}"
    assert None not in dfs.columns.names, f"None in dfs.columns.names={dfs.columns.names}"
    names = list(dfs.columns.names)
    dfs = dfs.unstack(key)
    keep_columns = dfs
    assert None not in dfs.index.names, f"None in dfs.index.names={dfs.index.names}"
    for n in names:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=FutureWarning)
            dfs = dfs.stack(n)
        assert (
            None not in dfs.index.names
        ), f"None in dfs.index.names={dfs.index.names}, n={n!r}"

    assert None not in dfs.index.names, f"None in dfs.index.names={dfs.index.names}"
    dfs = dfs.sort_index(axis=1).sort_index(axis=0)
    keep_columns = keep_columns.sort_index(axis=1).sort_index(axis=0)
    assert None not in dfs.index.names, f"None in dfs.index.names={dfs.index.names}"
    return dfs, keep_columns


def _reverse_column_names_order(
    df: pandas.DataFrame,  # :
    name: Optional[str] = None,
) -> pandas.DataFrame:  # :
    if len(df.columns.names) <= 1:
        return df
    col_names = df.columns.names
    assert isinstance(col_names, list), f"Unexpected type for {df.columns.names!r}"
    return df


def _select_metrics(
    df: pandas.DataFrame,
    select: List[Dict[str, str]],
    prefix: Optional[str] = None,
) -> pandas.DataFrame:
    suites = set(df.columns)
    rows = []
    names = list(df.index.names)
    set_names = set(names)
    for i in df.index.tolist():
        rows.append(set(dict(zip(names, i)).items()))

    subset = [
        (s, set({k: v for k, v in s.items() if k in set_names}.items())) for s in select
    ]

    keep = []
    for i, row in enumerate(rows):
        for j, (d, s) in enumerate(subset):
            if (s & row) == s:
                keep.append((i, d["new_name"], d["unit"], d.get("order", j), d["help"]))
                break

    dfi = df.iloc[[k[0] for k in keep]].reset_index(drop=False).copy()
    has_suite = "suite" in dfi.columns.names and "exporter" in dfi.columns.names

    def _mk(c):
        if not has_suite:
            return c
        cc = []
        for n in dfi.columns.names:
            if n == "exporter":
                cc.append(c)
            elif c in ("#order", "METRIC"):
                cc.append("")
            else:
                cc.append("~")
        return tuple(cc)

    dfi[_mk("METRIC")] = [k[1] for k in keep]
    dfi[_mk("#order")] = [k[3] for k in keep]
    dfi[_mk("unit")] = [k[2] for k in keep]
    dfi[_mk("~help")] = [k[4] for k in keep]
    dfi = dfi.copy()

    dd_ = set(select[0].keys())
    dd = set()
    for d in dd_:
        cs = [c for c in dfi.columns if d in c]
        if cs:
            dd.add(cs[0])
    cols = [
        *[c for c in dfi.columns if "#order" in c],
        *[c for c in dfi.columns if "METRIC" in c],
    ]
    skip = {*cols, *[c for c in dfi.columns if "unit" in c or "~help" in c]}
    for c in dfi.columns:
        if c in skip or c in dd:
            continue
        cols.append(c)
    cols.extend([c for c in dfi.columns if "unit" in c])
    cols.extend([c for c in dfi.columns if "~help" in c])
    dfi = dfi[cols].sort_values(cols[:-2])
    if prefix == "SIMPLE":
        # flat table
        import pandas

        col_suites = [c for c in dfi.columns if c in suites]
        dfs = []
        for cs in col_suites:
            to_drop = [c for c in col_suites if c != cs]
            if to_drop:
                df = dfi.drop(to_drop, axis=1).copy()
            else:
                df = dfi.copy()
            df["suite"] = cs
            df["value"] = df[cs]
            df = df.drop(cs, axis=1)
            df.columns = [str(c) for c in df.columns]
            dfs.append(df[~df["value"].isna()])
        dfi = pandas.concat(dfs, axis=0).reset_index(drop=True)
    return dfi.sort_index(axis=1), suites


def _filter_data(
    df: pandas.DataFrame,
    filter_in: Optional[str] = None,
    filter_out: Optional[str] = None,
) -> pandas.DataFrame:
    """
    Argument `filter` follows the syntax
    ``<column1>:<fmt1>/<column2>:<fmt2>``.

    The format is the following:

    * a value or a set of values separated by ``;``
    """
    if not filter_in and not filter_out:
        return df

    def _f(fmt):
        cond = {}
        if isinstance(fmt, str):
            cols = fmt.split("/")
            for c in cols:
                assert ":" in c, f"Unexpected value {c!r} in fmt={fmt!r}"
                spl = c.split(":")
                assert len(spl) == 2, f"Unexpected value {c!r} in fmt={fmt!r}"
                name, fil = spl
                cond[name] = FILTERS[fil] if fil in FILTERS else set(fil.split(";"))
        return cond

    if filter_in:
        cond = _f(filter_in)
        assert isinstance(cond, dict), f"Unexpected type {type(cond)} for fmt={filter_in!r}"
        for k, v in cond.items():
            if k not in df.columns:
                continue
            df = df[df[k].isin(v)]

    if filter_out:
        cond = _f(filter_out)
        assert isinstance(cond, dict), f"Unexpected type {type(cond)} for fmt={filter_out!r}"
        for k, v in cond.items():
            if k not in df.columns:
                continue
            df = df[~df[k].isin(v)]
    return df


def _select_model_metrics(
    res: Dict[str, pandas.DataFrame],
    select: List[Dict[str, str]],
    stack_levels: Sequence[str],
) -> pandas.DataFrame:
    import pandas

    concat = []
    for i, metric in enumerate(select):
        cat, stat, new_name, agg = (
            metric["cat"],
            metric["stat"],
            metric["new_name"],
            metric["agg"],
        )
        if new_name.startswith("average "):
            new_name = new_name[len("average ") :]
        if agg in {"TOTAL", "COUNT", "COUNT%", "MAX", "SUM", "NAN"}:
            continue
        name = f"{cat}_{stat}"
        if name not in res:
            continue
        df = res[name].copy()
        cols = list(df.columns)
        if len(cols) == 1:
            col = (cols[0],) if isinstance(cols[0], str) else tuple(cols[0])
            col = (i, cat, stat, new_name, *col)
            names = ["#order", "cat", "stat", "full_name", *df.columns.names]
            df.columns = pandas.MultiIndex.from_tuples([col], names=names)
            concat.append(df)
        else:
            cols = [((c,) if isinstance(c, str) else tuple(c)) for c in cols]
            cols = [(i, cat, stat, new_name, *c) for c in cols]
            names = ["#order", "cat", "stat", "full_name", *df.columns.names]
            df.columns = pandas.MultiIndex.from_tuples(cols, names=names)
            concat.append(df)
    df = pandas.concat(concat, axis=1)
    if stack_levels:
        for c in stack_levels:
            if c in df.columns.names:
                with warnings.catch_warnings():
                    warnings.simplefilter(
                        "ignore", category=(FutureWarning, PerformanceWarning)
                    )
                    df = df.stack(c, dropna=np.nan)
    df = df.sort_index(axis=1)
    return df
